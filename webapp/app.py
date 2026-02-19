from __future__ import annotations

from collections import deque
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional
import base64
import csv
import hmac
import json
import os
import queue
import secrets
import sqlite3
import subprocess
import sys
import threading
import time
import uuid
import urllib.error
import urllib.request
from urllib.parse import urlencode

try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover
    ZoneInfo = None  # type: ignore

try:
    from cryptography.fernet import Fernet
except Exception:  # pragma: no cover - fallback en entornos sin paquete
    Fernet = None  # type: ignore

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

ROOT_DIR = Path(__file__).resolve().parents[1]
STATIC_DIR = Path(__file__).resolve().parent / "static"
PYTHON_EXE = sys.executable
CONFIG_DIR = ROOT_DIR / "config"
DB_PATH = Path(os.getenv("TRADING_BOT_DB_PATH", str(CONFIG_DIR / "trading_bot.db"))).resolve()
NON_SIGNALS_CSV_PATH = ROOT_DIR / "Lector" / "data" / "non_signals.csv"
MT5_TERMINAL_DEFAULT = os.getenv("MT5_TERMINAL_PATH_DEFAULT", r"C:\Program Files\MetaTrader 5\terminal64.exe")
WEB_AUTH_FILE = CONFIG_DIR / "web_auth.json"
RUNTIME_ENV_KEY_FILE = CONFIG_DIR / "runtime_env.key"
RUNTIME_ENV_SCOPE = "runtime_env"
DEFAULT_CHANNELS = [
    ("TechnicalPips", "-1001287502434", ""),
    ("Metabear_Forex", "-1001422733304", ""),
    ("MiCanalPrueba", "-1002509518709", ""),
]
QUEUE_PENDING_DIR = ROOT_DIR / "queue" / "pending"

RESTART_TARGETS = {"operador", "lector", "both"}
ALERT_SEVERITY_RANK = {"info": 1, "warning": 2, "critical": 3}
PROCESS_NAME_LECTOR = "lector"
PROCESS_NAME_OPERADOR = "operador"
ALLOWED_PROFILE_CODES = {"SCALP", "SWING"}
DEFAULT_PROFILE_CODE = "SWING"
RUNTIME_ENV_KEYS = {
    PROCESS_NAME_LECTOR: [
        "PYTHONUNBUFFERED",
        "TRADING_BOT_DB_PATH",
        "TELEGRAM_API_ID",
        "TELEGRAM_API_HASH",
        "OPENAI_API_KEY",
        "OPENAI_MODEL",
        "OPENAI_BASE_URL",
    ],
    PROCESS_NAME_OPERADOR: [
        "PYTHONUNBUFFERED",
        "TRADING_BOT_DB_PATH",
        "MT5_TERMINAL_PATH",
        "MT5_LOGIN",
        "MT5_PASSWORD",
        "MT5_SERVER",
        "EXECUTION_MODE",
        "EXECUTION_PROFILE",
        "TOTAL_VOLUME",
        "NEAR_ENTRY_PIPS_MIN",
        "NEAR_ENTRY_SPREAD_MULT",
        "VERIFY_ORDER_AFTER_SEND",
        "AUTO_CLOSE_ON_MISMATCH",
    ],
}

DEFAULT_APP_SETTINGS: dict[str, str] = {
    "migration_profiles_scalp_swing_v2_done": "false",
    "auto_restart_enabled": "false",
    "auto_restart_interval_min": "240",
    "auto_restart_target": "operador",
    "auto_restart_next_at": "",
    "alerts_enabled": "true",
    "alerts_check_interval_sec": "10",
    "alerts_queue_pending_threshold": "50",
    "alerts_queue_oldest_sec": "180",
    "alerts_pending_order_sec": "1200",
    "alerts_error_window_min": "15",
    "alerts_error_count_threshold": "8",
    "alerts_no_tickets_threshold": "3",
    "alerts_drawdown_daily_usd": "-150",
    "alerts_stale_sync_sec": "60",
    "discord_enabled": "false",
    "discord_webhook_url": "",
    "discord_min_severity": "warning",
    "retention_enabled": "true",
    "retention_run_interval_min": "60",
    "retention_archive_enabled": "true",
    "retention_strategy_days": "180",
    "retention_operation_events_days": "180",
    "retention_telegram_messages_days": "365",
    "retention_alerts_events_days": "365",
    "retention_processed_events_days": "90",
    "retention_last_run_at": "",
}

app = FastAPI()
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
runtime_fernet = None
web_auth_credentials: dict[str, str] | None = None
URUGUAY_TZ = ZoneInfo("America/Montevideo") if ZoneInfo is not None else timezone(timedelta(hours=-3))


class LogHub:
    def __init__(self, max_lines: int = 800):
        self.buffer = deque(maxlen=max_lines)
        self.lock = threading.Lock()
        self.subscribers: set[queue.Queue[str]] = set()

    def publish(self, line: str | None) -> None:
        if line is None:
            return
        clean = line.rstrip("\r\n")
        with self.lock:
            self.buffer.append(clean)
            subs = list(self.subscribers)
        for q in subs:
            try:
                q.put_nowait(clean)
            except queue.Full:
                pass

    def subscribe(self):
        q: queue.Queue[str] = queue.Queue(maxsize=1000)
        with self.lock:
            self.subscribers.add(q)
            snapshot = list(self.buffer)
        return q, snapshot

    def unsubscribe(self, q: queue.Queue[str]):
        with self.lock:
            self.subscribers.discard(q)


class ProcessManager:
    def __init__(self, name: str, cmd: list[str], cwd: Path):
        self.name = name
        self.cmd = cmd
        self.cwd = cwd
        self.proc: subprocess.Popen[str] | None = None
        self.thread: threading.Thread | None = None
        self.log = LogHub()
        self.last_exit: int | None = None
        self.last_env: dict[str, str] | None = None

    def running(self) -> bool:
        return self.proc is not None and self.proc.poll() is None

    def start(self, env: dict) -> None:
        if self.running():
            raise RuntimeError(f"{self.name} already running")
        self.last_env = dict(env)
        self.log.publish(f"[{self.name}] starting...")
        self.proc = subprocess.Popen(
            self.cmd,
            cwd=self.cwd,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
        self.thread = threading.Thread(target=self._reader, daemon=True)
        self.thread.start()
        self.log.publish(f"[{self.name}] started pid={self.proc.pid}")

    def _reader(self) -> None:
        if not self.proc or not self.proc.stdout:
            return
        for line in self.proc.stdout:
            self.log.publish(line)
        rc = self.proc.poll()
        self.last_exit = rc
        self.log.publish(f"[{self.name}] exited code={rc}")
        try:
            self.proc.stdout.close()
        except Exception:
            pass

    def stop(self) -> None:
        if not self.running():
            return
        self.log.publish(f"[{self.name}] stopping...")
        assert self.proc is not None
        self.proc.terminate()
        try:
            self.proc.wait(timeout=10)
        except Exception:
            self.proc.kill()
        rc = self.proc.poll()
        self.last_exit = rc
        self.log.publish(f"[{self.name}] stopped code={rc}")
        self.proc = None

    def restart(self) -> None:
        if not self.last_env:
            raise RuntimeError(f"{self.name} has no previous env to restart")
        was_running = self.running()
        if was_running:
            self.stop()
            time.sleep(0.2)
        self.start(dict(self.last_env))


LECTOR_CMD = [PYTHON_EXE, "-u", str(ROOT_DIR / "Lector" / "main.py")]
OPERADOR_CMD = [PYTHON_EXE, "-u", str(ROOT_DIR / "Operador" / "daemon.py")]

lector_manager = ProcessManager("LECTOR", LECTOR_CMD, ROOT_DIR)
operador_manager = ProcessManager("OPERADOR", OPERADOR_CMD, ROOT_DIR)


class LectorStartRequest(BaseModel):
    telegram_api_id: int
    telegram_api_hash: str
    openai_api_key: str
    openai_model: str
    openai_base_url: Optional[str] = ""


class OperadorStartRequest(BaseModel):
    mt5_terminal_path: Optional[str] = ""
    mt5_login: Optional[int] = None
    mt5_password: str
    mt5_server: Optional[str] = ""
    total_volume: Optional[float] = None
    near_entry_pips_min: Optional[float] = None
    near_entry_spread_mult: Optional[float] = None
    verify_order_after_send: Optional[bool] = None
    auto_close_on_mismatch: Optional[bool] = None


class ExecutionProfilePayload(BaseModel):
    code: str
    name: str
    description: Optional[str] = ""
    is_system: Optional[bool] = False


class OperatorPresetPayload(BaseModel):
    name: str
    mt5_terminal_path: str
    mt5_login: int
    mt5_server: str
    execution_profile_id: Optional[int] = None
    total_volume: Optional[float] = 0.03
    near_entry_pips_min: Optional[float] = 1.0
    near_entry_spread_mult: Optional[float] = 2.0
    verify_order_after_send: Optional[bool] = True
    auto_close_on_mismatch: Optional[bool] = False
    is_default: Optional[bool] = False


class ChannelCreateRequest(BaseModel):
    name: str
    chat_id: str
    external_id: Optional[str] = ""
    is_active: bool = True


class ChannelUpdateRequest(BaseModel):
    name: str
    chat_id: str
    external_id: Optional[str] = ""
    is_active: bool = True


class AssignmentCreateRequest(BaseModel):
    channel_id: int
    config_id: int
    mode: str = "virtual"  # real|virtual
    is_active: bool = True


class AssignmentUpdateRequest(BaseModel):
    mode: str = "virtual"
    is_active: bool = True


class ChannelPresetSetActivePayload(BaseModel):
    is_active: bool


class RestartConfigPayload(BaseModel):
    enabled: bool
    interval_minutes: int
    target: str = "operador"


class RestartNowPayload(BaseModel):
    target: Optional[str] = None


class AlertsConfigPayload(BaseModel):
    alerts_enabled: bool = True
    alerts_check_interval_sec: int = 10
    alerts_queue_pending_threshold: int = 50
    alerts_queue_oldest_sec: int = 180
    alerts_pending_order_sec: int = 1200
    alerts_error_window_min: int = 15
    alerts_error_count_threshold: int = 8
    alerts_no_tickets_threshold: int = 3
    alerts_drawdown_daily_usd: float = -150.0
    alerts_stale_sync_sec: int = 60
    discord_enabled: bool = False
    discord_webhook_url: Optional[str] = ""
    discord_min_severity: str = "warning"


class OperationManualClosePayload(BaseModel):
    reason: Optional[str] = "Cerrada desde Panel web a mano"
    details: Optional[str] = ""
    close_in_mt5: Optional[bool] = False


class OperationsManualClosePayload(BaseModel):
    mode: Optional[str] = "all"  # all|real|virtual
    include_pending: Optional[bool] = True
    reason: Optional[str] = "Cerrada desde Panel web a mano"
    details: Optional[str] = ""
    close_in_mt5: Optional[bool] = False


class WebAuthPasswordPayload(BaseModel):
    password: str


def _db_conn() -> sqlite3.Connection:
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=30.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA busy_timeout=5000")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _column_names(conn: sqlite3.Connection, table_name: str) -> set[str]:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return {str(r[1]) for r in rows}


def _utc_now_iso() -> str:
    return datetime.now(URUGUAY_TZ).isoformat(timespec="seconds")


def _uy_from_epoch_iso(epoch_ts: float) -> str:
    return datetime.fromtimestamp(float(epoch_ts), URUGUAY_TZ).isoformat(timespec="seconds")


def _load_or_create_runtime_env_key() -> bytes:
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    if Fernet is None:
        return b""
    if RUNTIME_ENV_KEY_FILE.exists():
        key = RUNTIME_ENV_KEY_FILE.read_bytes().strip()
        if key:
            return key
    key = Fernet.generate_key()
    RUNTIME_ENV_KEY_FILE.write_bytes(key)
    return key


def _runtime_cipher():
    global runtime_fernet
    if Fernet is None:
        return None
    if runtime_fernet is None:
        runtime_fernet = Fernet(_load_or_create_runtime_env_key())
    return runtime_fernet


def _encrypt_runtime_payload(payload: dict) -> str:
    raw = json.dumps(payload, ensure_ascii=True, default=str).encode("utf-8")
    cipher = _runtime_cipher()
    if cipher is None:
        return "plain:" + base64.b64encode(raw).decode("ascii")
    token = cipher.encrypt(raw)
    return "enc:" + token.decode("ascii")


def _decrypt_runtime_payload(token: str) -> dict:
    if not token:
        return {}
    st = str(token or "")
    if st.startswith("plain:"):
        raw = base64.b64decode(st[6:]).decode("utf-8")
    else:
        cipher = _runtime_cipher()
        if cipher is None:
            return {}
        if st.startswith("enc:"):
            st = st[4:]
        raw = cipher.decrypt(st.encode("ascii")).decode("utf-8")
    obj = json.loads(raw)
    if isinstance(obj, dict):
        return obj
    return {}


def _load_or_create_web_auth_credentials() -> dict[str, str]:
    env_user = str(os.getenv("TRADING_BOT_WEB_USER", "") or "").strip()
    env_pass = str(os.getenv("TRADING_BOT_WEB_PASSWORD", "") or "").strip()
    if env_user and env_pass:
        return {"username": env_user, "password": env_pass, "source": "env"}

    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    if WEB_AUTH_FILE.exists():
        try:
            data = json.loads(WEB_AUTH_FILE.read_text(encoding="utf-8"))
            user = str(data.get("username", "") or "").strip()
            pwd = str(data.get("password", "") or "").strip()
            if user and pwd:
                return {"username": user, "password": pwd, "source": "file"}
        except Exception:
            pass

    user = "admin"
    pwd = secrets.token_urlsafe(20)
    WEB_AUTH_FILE.write_text(
        json.dumps({"username": user, "password": pwd, "created_at": _utc_now_iso()}, ensure_ascii=True, indent=2),
        encoding="utf-8",
    )
    return {"username": user, "password": pwd, "source": "generated"}


def _decode_basic_auth(header: str) -> tuple[str, str]:
    if not str(header).startswith("Basic "):
        return "", ""
    token = str(header)[6:].strip()
    if not token:
        return "", ""
    try:
        decoded = base64.b64decode(token).decode("utf-8")
    except Exception:
        return "", ""
    if ":" not in decoded:
        return "", ""
    user, pwd = decoded.split(":", 1)
    return str(user), str(pwd)


def _is_auth_valid(request: Request) -> bool:
    global web_auth_credentials
    if web_auth_credentials is None:
        web_auth_credentials = _load_or_create_web_auth_credentials()
        src = web_auth_credentials.get("source", "")
        user = web_auth_credentials.get("username", "")
        print(f"[SECURITY] Web auth activo. user='{user}' source={src} file='{WEB_AUTH_FILE}'.")
    expected_user = str(web_auth_credentials.get("username", ""))
    expected_pass = str(web_auth_credentials.get("password", ""))
    user, pwd = _decode_basic_auth(request.headers.get("authorization", ""))
    return hmac.compare_digest(user, expected_user) and hmac.compare_digest(pwd, expected_pass)


def _is_web_password_valid(password: str) -> bool:
    global web_auth_credentials
    if web_auth_credentials is None:
        web_auth_credentials = _load_or_create_web_auth_credentials()
    expected_pass = str(web_auth_credentials.get("password", ""))
    provided = str(password or "")
    return bool(expected_pass) and hmac.compare_digest(provided, expected_pass)


def _parse_iso_utc_or_none(value: str | None):
    s = str(value or "").strip()
    if not s:
        return None
    dt = None
    try:
        dt = datetime.fromisoformat(s)
    except Exception:
        pass
    if dt is None:
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            return None
    if dt.tzinfo is not None:
        return dt.astimezone(URUGUAY_TZ).replace(tzinfo=None)
    return dt


def _bool_from_str(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default
    s = str(value).strip().lower()
    if s in ("1", "true", "yes", "y", "on"):
        return True
    if s in ("0", "false", "no", "n", "off"):
        return False
    return default


def _int_from_str(value: str | None, default: int) -> int:
    try:
        return int(str(value or "").strip())
    except Exception:
        return int(default)


def _float_from_str(value: str | None, default: float) -> float:
    try:
        return float(str(value or "").strip())
    except Exception:
        return float(default)


def _setting_get_bool(values: dict[str, str], key: str, default: bool) -> bool:
    return _bool_from_str(values.get(key), default=default)


def _setting_get_int(values: dict[str, str], key: str, default: int) -> int:
    return _int_from_str(values.get(key), default=default)


def _setting_get_float(values: dict[str, str], key: str, default: float) -> float:
    return _float_from_str(values.get(key), default=default)


def _ensure_app_settings_table() -> None:
    with _db_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        now = _utc_now_iso()
        for key, value in DEFAULT_APP_SETTINGS.items():
            row = conn.execute("SELECT key FROM app_settings WHERE key = ?", (key,)).fetchone()
            if not row:
                conn.execute(
                    "INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, ?)",
                    (key, str(value), now),
                )
        conn.commit()


def _settings_values() -> dict[str, str]:
    _ensure_app_settings_table()
    values = dict(DEFAULT_APP_SETTINGS)
    with _db_conn() as conn:
        rows = conn.execute("SELECT key, value FROM app_settings").fetchall()
    for r in rows:
        values[str(r["key"])] = str(r["value"])
    return values


def _setting_set(key: str, value: str) -> None:
    _ensure_app_settings_table()
    now = _utc_now_iso()
    with _db_conn() as conn:
        conn.execute(
            """
            INSERT INTO app_settings (key, value, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
            """,
            (str(key), str(value), now),
        )
        conn.commit()


def _ensure_runtime_env_table() -> None:
    with _db_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS process_runtime_env (
                process_name TEXT PRIMARY KEY,
                env_blob TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.commit()


def _runtime_env_filter(process_name: str, env: dict[str, str]) -> dict[str, str]:
    keys = RUNTIME_ENV_KEYS.get(str(process_name), [])
    filtered: dict[str, str] = {}
    for k in keys:
        if k in env and str(env.get(k, "")).strip() != "":
            filtered[str(k)] = str(env[k])
    return filtered


def _persist_process_runtime_env(process_name: str, env: dict[str, str]) -> None:
    pname = str(process_name).strip().lower()
    if pname not in RUNTIME_ENV_KEYS:
        return
    _ensure_runtime_env_table()
    filtered = _runtime_env_filter(pname, env)
    if not filtered:
        return
    token = _encrypt_runtime_payload(filtered)
    now = _utc_now_iso()
    with _db_conn() as conn:
        conn.execute(
            """
            INSERT INTO process_runtime_env (process_name, env_blob, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(process_name) DO UPDATE SET
                env_blob = excluded.env_blob,
                updated_at = excluded.updated_at
            """,
            (pname, token, now),
        )
        conn.commit()


def _restore_process_runtime_env(process_name: str) -> dict[str, str]:
    pname = str(process_name).strip().lower()
    if pname not in RUNTIME_ENV_KEYS:
        return {}
    _ensure_runtime_env_table()
    with _db_conn() as conn:
        row = conn.execute(
            "SELECT env_blob FROM process_runtime_env WHERE process_name = ?",
            (pname,),
        ).fetchone()
    if not row:
        return {}
    try:
        payload = _decrypt_runtime_payload(str(row["env_blob"] or ""))
    except Exception:
        return {}
    filtered = _runtime_env_filter(pname, payload)
    base = os.environ.copy()
    base["PYTHONUNBUFFERED"] = "1"
    base["TRADING_BOT_DB_PATH"] = str(DB_PATH)
    base.update(filtered)
    return base


def _restore_process_manager_env_cache() -> None:
    if Fernet is None:
        print("[SECURITY] cryptography no está instalada: runtime env persistence usa fallback no cifrado.")
    lenv = _restore_process_runtime_env(PROCESS_NAME_LECTOR)
    if lenv:
        lector_manager.last_env = lenv
    oenv = _restore_process_runtime_env(PROCESS_NAME_OPERADOR)
    if oenv:
        operador_manager.last_env = oenv


def _queue_pending_stats() -> dict[str, int]:
    QUEUE_PENDING_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(QUEUE_PENDING_DIR.glob("*.json"))
    if not files:
        return {"count": 0, "oldest_age_sec": 0}
    now = time.time()
    oldest_mtime = min([f.stat().st_mtime for f in files])
    oldest_age = max(0, int(now - oldest_mtime))
    return {"count": len(files), "oldest_age_sec": oldest_age}


def _ensure_channels_table() -> None:
    with _db_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS telegram_channels (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                chat_id TEXT NOT NULL UNIQUE,
                external_id TEXT NOT NULL DEFAULT '',
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        cols = _column_names(conn, "telegram_channels")
        if "external_id" not in cols:
            conn.execute("ALTER TABLE telegram_channels ADD COLUMN external_id TEXT NOT NULL DEFAULT ''")
        if "is_active" not in cols:
            conn.execute("ALTER TABLE telegram_channels ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
        if "created_at" not in cols:
            conn.execute("ALTER TABLE telegram_channels ADD COLUMN created_at TEXT NOT NULL DEFAULT ''")
        if "updated_at" not in cols:
            conn.execute("ALTER TABLE telegram_channels ADD COLUMN updated_at TEXT NOT NULL DEFAULT ''")

        count = conn.execute("SELECT COUNT(1) FROM telegram_channels").fetchone()[0]
        if count == 0:
            now = _utc_now_iso()
            conn.executemany(
                """
                INSERT INTO telegram_channels (name, chat_id, external_id, is_active, created_at, updated_at)
                VALUES (?, ?, ?, 1, ?, ?)
                """,
                [(name, chat_id, ext_id or "", now, now) for name, chat_id, ext_id in DEFAULT_CHANNELS],
            )
        conn.commit()


def _ensure_execution_profiles_table() -> None:
    with _db_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS execution_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                is_system INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        now = _utc_now_iso()
        defaults = [
            ("SCALP", "Scalp", "Perfil rapido para señales cortas", 1),
            ("SWING", "Swing", "Perfil amplio para operaciones largas", 1),
        ]
        for code, name, description, is_system in defaults:
            row = conn.execute("SELECT id FROM execution_profiles WHERE code = ?", (code,)).fetchone()
            if not row:
                conn.execute(
                    """
                    INSERT INTO execution_profiles (code, name, description, is_system, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (code, name, description, int(is_system), now, now),
                )
            else:
                conn.execute(
                    """
                    UPDATE execution_profiles
                    SET name = ?, description = ?, is_system = 1, updated_at = ?
                    WHERE code = ?
                    """,
                    (name, description, now, code),
                )

        swing_row = conn.execute("SELECT id FROM execution_profiles WHERE code = 'SWING'").fetchone()
        swing_id = int(swing_row["id"]) if swing_row else 0
        if swing_id > 0:
            op_table = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='operator_presets'").fetchone()
            if op_table:
                conn.execute(
                    """
                    UPDATE operator_presets
                    SET execution_profile_id = ?
                    WHERE COALESCE(execution_profile_id, 0) <= 0
                       OR execution_profile_id IN (
                           SELECT id FROM execution_profiles WHERE UPPER(code) NOT IN ('SCALP','SWING')
                       )
                    """,
                    (swing_id,),
                )
            conn.execute("DELETE FROM execution_profiles WHERE UPPER(code) NOT IN ('SCALP','SWING')")
        conn.commit()


def _swing_profile_id(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT id FROM execution_profiles WHERE code = 'SWING'").fetchone()
    if row:
        return int(row["id"])
    now = _utc_now_iso()
    conn.execute(
        """
        INSERT INTO execution_profiles (code, name, description, is_system, created_at, updated_at)
        VALUES ('SWING', 'Swing', 'Perfil amplio para operaciones largas', 1, ?, ?)
        """,
        (now, now),
    )
    rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return int(rid)


def _ensure_operator_presets_table() -> None:
    _ensure_execution_profiles_table()
    with _db_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS operator_presets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                mt5_terminal_path TEXT NOT NULL,
                mt5_login INTEGER NOT NULL,
                mt5_server TEXT NOT NULL,
                execution_profile_id INTEGER NOT NULL DEFAULT 0,
                total_volume REAL NOT NULL DEFAULT 0.03,
                near_entry_pips_min REAL NOT NULL DEFAULT 1.0,
                near_entry_spread_mult REAL NOT NULL DEFAULT 2.0,
                verify_order_after_send INTEGER NOT NULL DEFAULT 1,
                auto_close_on_mismatch INTEGER NOT NULL DEFAULT 0,
                is_default INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        cols = _column_names(conn, "operator_presets")
        if "is_default" not in cols:
            conn.execute("ALTER TABLE operator_presets ADD COLUMN is_default INTEGER NOT NULL DEFAULT 0")
        if "execution_profile_id" not in cols:
            conn.execute("ALTER TABLE operator_presets ADD COLUMN execution_profile_id INTEGER NOT NULL DEFAULT 0")
        if "total_volume" not in cols:
            conn.execute("ALTER TABLE operator_presets ADD COLUMN total_volume REAL NOT NULL DEFAULT 0.03")
        if "near_entry_pips_min" not in cols:
            conn.execute("ALTER TABLE operator_presets ADD COLUMN near_entry_pips_min REAL NOT NULL DEFAULT 1.0")
        if "near_entry_spread_mult" not in cols:
            conn.execute("ALTER TABLE operator_presets ADD COLUMN near_entry_spread_mult REAL NOT NULL DEFAULT 2.0")
        if "verify_order_after_send" not in cols:
            conn.execute("ALTER TABLE operator_presets ADD COLUMN verify_order_after_send INTEGER NOT NULL DEFAULT 1")
        if "auto_close_on_mismatch" not in cols:
            conn.execute("ALTER TABLE operator_presets ADD COLUMN auto_close_on_mismatch INTEGER NOT NULL DEFAULT 0")
        if "created_at" not in cols:
            conn.execute("ALTER TABLE operator_presets ADD COLUMN created_at TEXT NOT NULL DEFAULT ''")
        if "updated_at" not in cols:
            conn.execute("ALTER TABLE operator_presets ADD COLUMN updated_at TEXT NOT NULL DEFAULT ''")
        swing_id = _swing_profile_id(conn)
        conn.execute(
            """
            UPDATE operator_presets
            SET execution_profile_id = ?
            WHERE COALESCE(execution_profile_id, 0) <= 0
               OR execution_profile_id IN (
                   SELECT id FROM execution_profiles WHERE UPPER(code) NOT IN ('SCALP','SWING')
               )
            """,
            (swing_id,),
        )
        conn.commit()


def _ensure_assignments_table() -> None:
    with _db_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS channel_config_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                channel_id INTEGER NOT NULL,
                config_id INTEGER NOT NULL,
                mode TEXT NOT NULL CHECK (mode IN ('real','virtual')),
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(channel_id, config_id)
            )
            """
        )
        cols = _column_names(conn, "channel_config_assignments")
        if "mode" not in cols:
            conn.execute("ALTER TABLE channel_config_assignments ADD COLUMN mode TEXT NOT NULL DEFAULT 'virtual'")
        if "is_active" not in cols:
            conn.execute("ALTER TABLE channel_config_assignments ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
        if "created_at" not in cols:
            conn.execute("ALTER TABLE channel_config_assignments ADD COLUMN created_at TEXT NOT NULL DEFAULT ''")
        if "updated_at" not in cols:
            conn.execute("ALTER TABLE channel_config_assignments ADD COLUMN updated_at TEXT NOT NULL DEFAULT ''")

        # Garantiza máximo 1 asignación en modo real por canal.
        now = _utc_now_iso()
        channel_rows = conn.execute("SELECT DISTINCT channel_id FROM channel_config_assignments").fetchall()
        for ch in channel_rows:
            cid = int(ch["channel_id"])
            real_rows = conn.execute(
                """
                SELECT id
                FROM channel_config_assignments
                WHERE channel_id = ? AND mode = 'real'
                ORDER BY is_active DESC, id ASC
                """,
                (cid,),
            ).fetchall()
            if len(real_rows) <= 1:
                continue
            keep_id = int(real_rows[0]["id"])
            conn.execute(
                """
                UPDATE channel_config_assignments
                SET mode = 'virtual', updated_at = ?
                WHERE channel_id = ? AND mode = 'real' AND id != ?
                """,
                (now, cid, keep_id),
            )

        # Real solo permitido con preset de perfil SWING.
        conn.execute(
            """
            UPDATE channel_config_assignments
            SET mode = 'virtual', updated_at = ?
            WHERE mode = 'real'
              AND config_id IN (
                  SELECT p.id
                  FROM operator_presets p
                  LEFT JOIN execution_profiles ep ON ep.id = p.execution_profile_id
                  WHERE UPPER(COALESCE(ep.code, 'SWING')) != 'SWING'
              )
            """,
            (now,),
        )
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_assignments_one_real_per_channel
            ON channel_config_assignments(channel_id)
            WHERE mode = 'real'
            """
        )
        conn.commit()


def _ensure_channel_preset_events_table() -> None:
    with _db_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS channel_preset_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL,
                assignment_id INTEGER NOT NULL,
                event_type TEXT NOT NULL,
                channel_id INTEGER,
                channel_name TEXT,
                channel_chat_id TEXT,
                config_id INTEGER,
                config_name TEXT,
                mode TEXT,
                is_active INTEGER NOT NULL DEFAULT 0,
                details TEXT NOT NULL DEFAULT '',
                metadata_json TEXT NOT NULL DEFAULT ''
            )
            """
        )
        cols = _column_names(conn, "channel_preset_events")
        missing = {
            "ts": "TEXT NOT NULL DEFAULT ''",
            "assignment_id": "INTEGER NOT NULL DEFAULT 0",
            "event_type": "TEXT NOT NULL DEFAULT ''",
            "channel_id": "INTEGER",
            "channel_name": "TEXT",
            "channel_chat_id": "TEXT",
            "config_id": "INTEGER",
            "config_name": "TEXT",
            "mode": "TEXT",
            "is_active": "INTEGER NOT NULL DEFAULT 0",
            "details": "TEXT NOT NULL DEFAULT ''",
            "metadata_json": "TEXT NOT NULL DEFAULT ''",
        }
        for col, col_sql in missing.items():
            if col not in cols:
                conn.execute(f"ALTER TABLE channel_preset_events ADD COLUMN {col} {col_sql}")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_cp_events_assignment_ts ON channel_preset_events(assignment_id, ts)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_cp_events_ts ON channel_preset_events(ts)")
        conn.commit()


def _assignment_snapshot_by_id(conn: sqlite3.Connection, assignment_id: int):
    return conn.execute(
        """
        SELECT
            a.id AS assignment_id, a.channel_id, a.config_id, a.mode, a.is_active,
            c.name AS channel_name, c.chat_id AS channel_chat_id,
            p.name AS config_name
        FROM channel_config_assignments a
        LEFT JOIN telegram_channels c ON c.id = a.channel_id
        LEFT JOIN operator_presets p ON p.id = a.config_id
        WHERE a.id = ?
        """,
        (int(assignment_id),),
    ).fetchone()


def _assignment_row_by_id(conn: sqlite3.Connection, assignment_id: int):
    return conn.execute(
        """
        SELECT
            a.id, a.channel_id, a.config_id, a.mode, a.is_active, a.created_at, a.updated_at,
            c.name AS channel_name, c.chat_id AS channel_chat_id,
            p.name AS config_name
        FROM channel_config_assignments a
        LEFT JOIN telegram_channels c ON c.id = a.channel_id
        LEFT JOIN operator_presets p ON p.id = a.config_id
        WHERE a.id = ?
        """,
        (int(assignment_id),),
    ).fetchone()


def _open_operations_count_for_channel_preset_conn(conn: sqlite3.Connection, channel_id: int, config_id: int) -> int:
    row = conn.execute(
        """
        SELECT COUNT(1) AS c
        FROM operation_records
        WHERE channel_id = ? AND preset_id = ? AND status IN ('OPEN','PENDING')
        """,
        (int(channel_id), int(config_id)),
    ).fetchone()
    return int(row["c"] if row is not None else 0)


def _append_channel_preset_event_conn(
    conn: sqlite3.Connection,
    *,
    assignment_id: int,
    event_type: str,
    snapshot,
    details: str = "",
    metadata: dict | None = None,
    ts: str | None = None,
):
    now = str(ts or _utc_now_iso())
    meta_json = json.dumps(metadata or {}, ensure_ascii=True, default=str)
    conn.execute(
        """
        INSERT INTO channel_preset_events (
            ts, assignment_id, event_type,
            channel_id, channel_name, channel_chat_id,
            config_id, config_name, mode, is_active, details, metadata_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            now,
            int(assignment_id),
            str(event_type),
            int(snapshot["channel_id"]) if snapshot and snapshot["channel_id"] is not None else None,
            str(snapshot["channel_name"] or "") if snapshot else "",
            str(snapshot["channel_chat_id"] or "") if snapshot else "",
            int(snapshot["config_id"]) if snapshot and snapshot["config_id"] is not None else None,
            str(snapshot["config_name"] or "") if snapshot else "",
            str(snapshot["mode"] or "") if snapshot else "",
            int(snapshot["is_active"] or 0) if snapshot else 0,
            str(details or ""),
            meta_json,
        ),
    )


def _seed_channel_preset_events_for_existing_assignments() -> None:
    _ensure_channel_preset_events_table()
    _ensure_assignments_table()
    now = _utc_now_iso()
    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT a.id
            FROM channel_config_assignments a
            WHERE NOT EXISTS (
                SELECT 1
                FROM channel_preset_events e
                WHERE e.assignment_id = a.id
            )
            ORDER BY a.id ASC
            """
        ).fetchall()
        for r in rows:
            aid = int(r["id"])
            snap = _assignment_snapshot_by_id(conn, aid)
            _append_channel_preset_event_conn(
                conn,
                assignment_id=aid,
                event_type="bootstrap",
                snapshot=snap,
                details="bootstrap_existing_assignment",
                metadata={"source": "startup"},
                ts=now,
            )
        conn.commit()


def _ensure_reports_tables() -> None:
    with _db_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS strategy_event_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL,
                event_id TEXT,
                message_id TEXT,
                channel_id INTEGER,
                channel_name TEXT,
                config_id INTEGER,
                config_name TEXT,
                mode TEXT,
                event_type TEXT,
                symbol TEXT,
                side TEXT,
                operator_class TEXT,
                entry_message_id TEXT,
                reply_to TEXT,
                status TEXT,
                error_type TEXT,
                pnl_usd REAL,
                pnl_pips REAL,
                details TEXT
            )
            """
        )
        cols = _column_names(conn, "strategy_event_log")
        if "message_id" not in cols:
            conn.execute("ALTER TABLE strategy_event_log ADD COLUMN message_id TEXT")
        if "operator_class" not in cols:
            conn.execute("ALTER TABLE strategy_event_log ADD COLUMN operator_class TEXT")

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS operation_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                operation_key TEXT NOT NULL UNIQUE,
                mode TEXT NOT NULL,
                status TEXT NOT NULL,
                is_virtual INTEGER NOT NULL DEFAULT 0,
                channel_id INTEGER,
                channel_name TEXT,
                channel_index TEXT,
                preset_id INTEGER,
                preset_name TEXT,
                execution_profile TEXT,
                symbol TEXT,
                side TEXT,
                entry_message_id TEXT,
                entry_event_id TEXT,
                entry_trigger_message_id TEXT,
                entry_ts TEXT,
                opened_at TEXT,
                ticket TEXT,
                comment TEXT,
                volume REAL,
                entry_price REAL,
                sl REAL,
                tp REAL,
                had_modifications INTEGER NOT NULL DEFAULT 0,
                modifications_count INTEGER NOT NULL DEFAULT 0,
                last_modification_message_id TEXT,
                last_modified_at TEXT,
                closed_at TEXT,
                close_event_id TEXT,
                close_trigger_message_id TEXT,
                close_reason TEXT,
                close_source TEXT,
                close_error_id TEXT,
                close_error_type TEXT,
                close_details TEXT,
                pnl_usd REAL,
                pnl_pips REAL,
                last_pips REAL,
                last_profit_usd REAL,
                duration_seconds INTEGER,
                last_sync_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        op_cols = _column_names(conn, "operation_records")
        op_missing_sql = {
            "operation_key": "TEXT NOT NULL DEFAULT ''",
            "mode": "TEXT NOT NULL DEFAULT ''",
            "status": "TEXT NOT NULL DEFAULT ''",
            "is_virtual": "INTEGER NOT NULL DEFAULT 0",
            "channel_id": "INTEGER",
            "channel_name": "TEXT",
            "channel_index": "TEXT",
            "preset_id": "INTEGER",
            "preset_name": "TEXT",
            "execution_profile": "TEXT",
            "symbol": "TEXT",
            "side": "TEXT",
            "entry_message_id": "TEXT",
            "entry_event_id": "TEXT",
            "entry_trigger_message_id": "TEXT",
            "entry_ts": "TEXT",
            "opened_at": "TEXT",
            "ticket": "TEXT",
            "comment": "TEXT",
            "volume": "REAL",
            "entry_price": "REAL",
            "sl": "REAL",
            "tp": "REAL",
            "had_modifications": "INTEGER NOT NULL DEFAULT 0",
            "modifications_count": "INTEGER NOT NULL DEFAULT 0",
            "last_modification_message_id": "TEXT",
            "last_modified_at": "TEXT",
            "closed_at": "TEXT",
            "close_event_id": "TEXT",
            "close_trigger_message_id": "TEXT",
            "close_reason": "TEXT",
            "close_source": "TEXT",
            "close_error_id": "TEXT",
            "close_error_type": "TEXT",
            "close_details": "TEXT",
            "pnl_usd": "REAL",
            "pnl_pips": "REAL",
            "last_pips": "REAL",
            "last_profit_usd": "REAL",
            "duration_seconds": "INTEGER",
            "last_sync_at": "TEXT",
            "created_at": "TEXT NOT NULL DEFAULT ''",
            "updated_at": "TEXT NOT NULL DEFAULT ''",
        }
        for col, col_sql in op_missing_sql.items():
            if col not in op_cols:
                conn.execute(f"ALTER TABLE operation_records ADD COLUMN {col} {col_sql}")

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS operation_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                operation_id INTEGER NOT NULL,
                ts TEXT NOT NULL,
                event_type TEXT NOT NULL,
                event_id TEXT,
                message_id TEXT,
                reply_to TEXT,
                status TEXT,
                error_type TEXT,
                sl REAL,
                tp REAL,
                pnl_usd REAL,
                pnl_pips REAL,
                details TEXT
            )
            """
        )
        op_ev_cols = _column_names(conn, "operation_events")
        op_ev_missing = {
            "operation_id": "INTEGER NOT NULL DEFAULT 0",
            "ts": "TEXT NOT NULL DEFAULT ''",
            "event_type": "TEXT NOT NULL DEFAULT ''",
            "event_id": "TEXT",
            "message_id": "TEXT",
            "reply_to": "TEXT",
            "status": "TEXT",
            "error_type": "TEXT",
            "sl": "REAL",
            "tp": "REAL",
            "pnl_usd": "REAL",
            "pnl_pips": "REAL",
            "details": "TEXT",
        }
        for col, col_sql in op_ev_missing.items():
            if col not in op_ev_cols:
                conn.execute(f"ALTER TABLE operation_events ADD COLUMN {col} {col_sql}")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS telegram_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                message_uid TEXT NOT NULL,
                message_key TEXT NOT NULL DEFAULT '',
                message_id TEXT NOT NULL,
                channel_name TEXT,
                channel_index TEXT,
                ts TEXT,
                reply_to TEXT,
                event_id TEXT,
                event_type TEXT,
                symbol TEXT,
                operation TEXT,
                operator_class TEXT,
                message_text TEXT,
                raw_payload TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        tmsg_cols = _column_names(conn, "telegram_messages")
        tmsg_missing = {
            "message_uid": "TEXT NOT NULL DEFAULT ''",
            "message_key": "TEXT NOT NULL DEFAULT ''",
            "message_id": "TEXT NOT NULL DEFAULT ''",
            "channel_name": "TEXT",
            "channel_index": "TEXT",
            "ts": "TEXT",
            "reply_to": "TEXT",
            "event_id": "TEXT",
            "event_type": "TEXT",
            "symbol": "TEXT",
            "operation": "TEXT",
            "operator_class": "TEXT",
            "message_text": "TEXT",
            "raw_payload": "TEXT",
            "created_at": "TEXT NOT NULL DEFAULT ''",
            "updated_at": "TEXT NOT NULL DEFAULT ''",
        }
        for col, col_sql in tmsg_missing.items():
            if col not in tmsg_cols:
                conn.execute(f"ALTER TABLE telegram_messages ADD COLUMN {col} {col_sql}")
        conn.execute(
            """
            UPDATE telegram_messages
            SET message_key = CASE
                WHEN COALESCE(message_key, '') = '' AND COALESCE(channel_name, '') <> '' AND COALESCE(message_id, '') <> ''
                THEN channel_name || ':' || message_id
                WHEN COALESCE(message_key, '') = '' THEN COALESCE(message_uid, '')
                ELSE message_key
            END
            WHERE COALESCE(message_key, '') = ''
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS alerts_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL,
                source TEXT NOT NULL,
                code TEXT NOT NULL,
                severity TEXT NOT NULL,
                title TEXT NOT NULL,
                details TEXT NOT NULL,
                data_json TEXT NOT NULL DEFAULT '',
                is_active INTEGER NOT NULL DEFAULT 1,
                first_seen TEXT NOT NULL,
                last_seen TEXT NOT NULL,
                resolved_at TEXT,
                occurrences INTEGER NOT NULL DEFAULT 1,
                notified_discord INTEGER NOT NULL DEFAULT 0
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS alerts_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL,
                alert_id INTEGER,
                source TEXT NOT NULL,
                code TEXT NOT NULL,
                severity TEXT NOT NULL,
                title TEXT NOT NULL,
                details TEXT NOT NULL,
                data_json TEXT NOT NULL DEFAULT '',
                event_type TEXT NOT NULL DEFAULT 'active',
                status TEXT NOT NULL DEFAULT 'ACTIVA',
                is_active INTEGER NOT NULL DEFAULT 1,
                occurrences INTEGER NOT NULL DEFAULT 1,
                first_seen TEXT,
                last_seen TEXT,
                resolved_at TEXT
            )
            """
        )
        alert_cols = _column_names(conn, "alerts_log")
        alert_missing = {
            "source": "TEXT NOT NULL DEFAULT ''",
            "code": "TEXT NOT NULL DEFAULT ''",
            "severity": "TEXT NOT NULL DEFAULT 'warning'",
            "title": "TEXT NOT NULL DEFAULT ''",
            "details": "TEXT NOT NULL DEFAULT ''",
            "data_json": "TEXT NOT NULL DEFAULT ''",
            "is_active": "INTEGER NOT NULL DEFAULT 1",
            "first_seen": "TEXT NOT NULL DEFAULT ''",
            "last_seen": "TEXT NOT NULL DEFAULT ''",
            "resolved_at": "TEXT",
            "occurrences": "INTEGER NOT NULL DEFAULT 1",
            "notified_discord": "INTEGER NOT NULL DEFAULT 0",
        }
        for col, col_sql in alert_missing.items():
            if col not in alert_cols:
                conn.execute(f"ALTER TABLE alerts_log ADD COLUMN {col} {col_sql}")
        alert_event_cols = _column_names(conn, "alerts_events")
        alert_event_missing = {
            "alert_id": "INTEGER",
            "source": "TEXT NOT NULL DEFAULT ''",
            "code": "TEXT NOT NULL DEFAULT ''",
            "severity": "TEXT NOT NULL DEFAULT 'warning'",
            "title": "TEXT NOT NULL DEFAULT ''",
            "details": "TEXT NOT NULL DEFAULT ''",
            "data_json": "TEXT NOT NULL DEFAULT ''",
            "event_type": "TEXT NOT NULL DEFAULT 'active'",
            "status": "TEXT NOT NULL DEFAULT 'ACTIVA'",
            "is_active": "INTEGER NOT NULL DEFAULT 1",
            "occurrences": "INTEGER NOT NULL DEFAULT 1",
            "first_seen": "TEXT",
            "last_seen": "TEXT",
            "resolved_at": "TEXT",
        }
        for col, col_sql in alert_event_missing.items():
            if col not in alert_event_cols:
                conn.execute(f"ALTER TABLE alerts_events ADD COLUMN {col} {col_sql}")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_strategy_log_ts ON strategy_event_log(ts)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_strategy_log_combo ON strategy_event_log(channel_id, config_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_op_records_state ON operation_records(status, mode, opened_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_op_records_entry ON operation_records(channel_index, entry_message_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_op_events_operation ON operation_events(operation_id, ts)")
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_tmsg_uid ON telegram_messages(message_uid)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_tmsg_message_key ON telegram_messages(message_key, ts)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_tmsg_message_id ON telegram_messages(message_id, ts)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_alerts_active_code ON alerts_log(is_active, code)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_alerts_ts ON alerts_log(ts)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_alerts_events_ts ON alerts_events(ts)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_alerts_events_code ON alerts_events(code, ts)")
        conn.commit()


def _normalize_operator_preset_payload(payload: OperatorPresetPayload):
    name = str(payload.name or "").strip()
    terminal = str(payload.mt5_terminal_path or "").strip()
    server = str(payload.mt5_server or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Preset name is required")
    if not terminal:
        raise HTTPException(status_code=400, detail="MT5 terminal path is required")
    if not server:
        raise HTTPException(status_code=400, detail="MT5 server is required")
    if int(payload.mt5_login) <= 0:
        raise HTTPException(status_code=400, detail="MT5 login must be a positive integer")
    total_volume = float(payload.total_volume if payload.total_volume is not None else 0.03)
    near_min = float(payload.near_entry_pips_min if payload.near_entry_pips_min is not None else 1.0)
    near_mult = float(payload.near_entry_spread_mult if payload.near_entry_spread_mult is not None else 2.0)
    if total_volume <= 0:
        raise HTTPException(status_code=400, detail="TOTAL_VOLUME must be greater than 0")
    if near_min < 0 or near_mult < 0:
        raise HTTPException(status_code=400, detail="NEAR_ENTRY params cannot be negative")
    profile_id = int(payload.execution_profile_id) if payload.execution_profile_id is not None else 0
    if profile_id <= 0:
        raise HTTPException(status_code=400, detail="execution_profile_id is required")
    return {
        "name": name,
        "mt5_terminal_path": terminal,
        "mt5_login": int(payload.mt5_login),
        "mt5_server": server,
        "execution_profile_id": profile_id,
        "total_volume": total_volume,
        "near_entry_pips_min": near_min,
        "near_entry_spread_mult": near_mult,
        "verify_order_after_send": 1 if bool(payload.verify_order_after_send) else 0,
        "auto_close_on_mismatch": 1 if bool(payload.auto_close_on_mismatch) else 0,
        "is_default": 1 if bool(payload.is_default) else 0,
    }


def _list_operator_presets() -> list[dict]:
    _ensure_operator_presets_table()
    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                p.id, p.name, p.mt5_terminal_path, p.mt5_login, p.mt5_server,
                p.execution_profile_id,
                COALESCE(ep.code, 'SWING') AS execution_profile_code,
                COALESCE(ep.name, 'Swing') AS execution_profile_name,
                p.total_volume, p.near_entry_pips_min, p.near_entry_spread_mult,
                p.verify_order_after_send, p.auto_close_on_mismatch, p.is_default, p.created_at, p.updated_at
            FROM operator_presets p
            LEFT JOIN execution_profiles ep ON ep.id = p.execution_profile_id
            ORDER BY p.is_default DESC, p.id ASC
            """
        ).fetchall()
    return [
        {
            "id": int(r["id"]),
            "name": str(r["name"]),
            "mt5_terminal_path": str(r["mt5_terminal_path"]),
            "mt5_login": int(r["mt5_login"]),
            "mt5_server": str(r["mt5_server"]),
            "execution_profile_id": int(r["execution_profile_id"]),
            "execution_profile_code": str(r["execution_profile_code"]),
            "execution_profile_name": str(r["execution_profile_name"]),
            "total_volume": float(r["total_volume"]),
            "near_entry_pips_min": float(r["near_entry_pips_min"]),
            "near_entry_spread_mult": float(r["near_entry_spread_mult"]),
            "verify_order_after_send": bool(r["verify_order_after_send"]),
            "auto_close_on_mismatch": bool(r["auto_close_on_mismatch"]),
            "is_default": bool(r["is_default"]),
            "created_at": str(r["created_at"] or ""),
            "updated_at": str(r["updated_at"] or ""),
        }
        for r in rows
    ]


def _normalize_assignment_mode(mode: str) -> str:
    m = str(mode or "").strip().lower()
    return "real" if m == "real" else "virtual"


def _preset_profile_code(conn: sqlite3.Connection, preset_id: int) -> str:
    row = conn.execute(
        """
        SELECT COALESCE(ep.code, ?) AS code
        FROM operator_presets p
        LEFT JOIN execution_profiles ep ON ep.id = p.execution_profile_id
        WHERE p.id = ?
        """,
        (DEFAULT_PROFILE_CODE, int(preset_id)),
    ).fetchone()
    if not row:
        return DEFAULT_PROFILE_CODE
    return str(row["code"] or DEFAULT_PROFILE_CODE).strip().upper() or DEFAULT_PROFILE_CODE


def _normalize_execution_profile_payload(payload: ExecutionProfilePayload) -> dict:
    code = str(payload.code or "").strip().upper()
    name = str(payload.name or "").strip()
    description = str(payload.description or "").strip()
    if not code:
        raise HTTPException(status_code=400, detail="Profile code is required")
    if not name:
        raise HTTPException(status_code=400, detail="Profile name is required")
    if len(code) > 24:
        raise HTTPException(status_code=400, detail="Profile code too long (max 24)")
    if not all(ch.isalnum() or ch in {"_", "-"} for ch in code):
        raise HTTPException(status_code=400, detail="Profile code solo permite letras, números, _ y -")
    return {
        "code": code,
        "name": name,
        "description": description,
        "is_system": 1 if bool(payload.is_system) else 0,
    }


def _list_execution_profiles() -> list[dict]:
    _ensure_execution_profiles_table()
    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT id, code, name, description, is_system, created_at, updated_at
            FROM execution_profiles
            ORDER BY is_system DESC, code ASC, id ASC
            """
        ).fetchall()
    return [
        {
            "id": int(r["id"]),
            "code": str(r["code"]),
            "name": str(r["name"]),
            "description": str(r["description"] or ""),
            "is_system": bool(r["is_system"]),
            "created_at": str(r["created_at"] or ""),
            "updated_at": str(r["updated_at"] or ""),
        }
        for r in rows
    ]


def _list_assignments() -> list[dict]:
    _ensure_assignments_table()
    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                a.id, a.channel_id, a.config_id, a.mode, a.is_active, a.created_at, a.updated_at,
                c.name AS channel_name, c.chat_id AS channel_chat_id,
                p.name AS config_name, p.execution_profile_id,
                COALESCE(ep.code, 'SWING') AS execution_profile_code,
                COALESCE(ep.name, 'Swing') AS execution_profile_name,
                p.total_volume,
                p.near_entry_pips_min, p.near_entry_spread_mult,
                p.verify_order_after_send, p.auto_close_on_mismatch
            FROM channel_config_assignments a
            JOIN telegram_channels c ON c.id = a.channel_id
            JOIN operator_presets p ON p.id = a.config_id
            LEFT JOIN execution_profiles ep ON ep.id = p.execution_profile_id
            ORDER BY c.name ASC, a.mode ASC, p.name ASC
            """
        ).fetchall()
    return [
        {
            "id": int(r["id"]),
            "channel_id": int(r["channel_id"]),
            "channel_name": str(r["channel_name"]),
            "channel_chat_id": str(r["channel_chat_id"]),
            "config_id": int(r["config_id"]),
            "config_name": str(r["config_name"]),
            "preset_id": int(r["config_id"]),
            "preset_name": str(r["config_name"]),
            "mode": str(r["mode"]),
            "is_active": bool(r["is_active"]),
            "execution_profile_id": int(r["execution_profile_id"]),
            "execution_profile_code": str(r["execution_profile_code"]),
            "execution_profile_name": str(r["execution_profile_name"]),
            "total_volume": float(r["total_volume"]),
            "near_entry_pips_min": float(r["near_entry_pips_min"]),
            "near_entry_spread_mult": float(r["near_entry_spread_mult"]),
            "verify_order_after_send": bool(r["verify_order_after_send"]),
            "auto_close_on_mismatch": bool(r["auto_close_on_mismatch"]),
            "created_at": str(r["created_at"] or ""),
            "updated_at": str(r["updated_at"] or ""),
        }
        for r in rows
    ]


def _select_auto_real_preset_id_conn(conn: sqlite3.Connection) -> int | None:
    rows = conn.execute(
        """
        SELECT
            p.id,
            p.is_default,
            UPPER(COALESCE(ep.code, ?)) AS execution_profile_code
        FROM operator_presets p
        LEFT JOIN execution_profiles ep ON ep.id = p.execution_profile_id
        ORDER BY p.is_default DESC, p.id ASC
        """,
        (DEFAULT_PROFILE_CODE,),
    ).fetchall()
    swing_ids = [int(r["id"]) for r in rows if str(r["execution_profile_code"] or DEFAULT_PROFILE_CODE).upper() == "SWING"]
    if not swing_ids:
        return None
    for r in rows:
        if int(r["is_default"] or 0) == 1 and int(r["id"]) in swing_ids:
            return int(r["id"])
    existing = conn.execute(
        """
        SELECT config_id, COUNT(1) AS cnt
        FROM channel_config_assignments
        WHERE mode = 'real'
        GROUP BY config_id
        ORDER BY cnt DESC, config_id ASC
        LIMIT 1
        """
    ).fetchone()
    if existing is not None:
        cid = int(existing["config_id"])
        if cid in swing_ids:
            return cid
    return int(sorted(swing_ids)[0])


def _sync_channel_preset_assignments_conn(
    conn: sqlite3.Connection,
    *,
    reason: str = "auto_sync",
    metadata: dict | None = None,
    now_ts: str | None = None,
) -> dict:
    now = str(now_ts or _utc_now_iso())
    meta = dict(metadata or {})
    channels = conn.execute(
        """
        SELECT id
        FROM telegram_channels
        ORDER BY id ASC
        """
    ).fetchall()
    presets = conn.execute(
        """
        SELECT id
        FROM operator_presets
        ORDER BY is_default DESC, id ASC
        """
    ).fetchall()
    existing_rows = conn.execute(
        """
        SELECT id, channel_id, config_id, mode, is_active
        FROM channel_config_assignments
        ORDER BY id ASC
        """
    ).fetchall()

    channel_ids = [int(r["id"]) for r in channels]
    preset_ids = [int(r["id"]) for r in presets]
    desired_pairs = {(int(cid), int(pid)) for cid in channel_ids for pid in preset_ids}
    existing_by_pair: dict[tuple[int, int], sqlite3.Row] = {}
    created = 0
    updated = 0
    deleted = 0

    for row in existing_rows:
        key = (int(row["channel_id"]), int(row["config_id"]))
        if key not in desired_pairs:
            delete_snap = _assignment_snapshot_by_id(conn, int(row["id"]))
            if delete_snap:
                delete_snap = {
                    "channel_id": delete_snap["channel_id"],
                    "channel_name": delete_snap["channel_name"],
                    "channel_chat_id": delete_snap["channel_chat_id"],
                    "config_id": delete_snap["config_id"],
                    "config_name": delete_snap["config_name"],
                    "mode": delete_snap["mode"],
                    "is_active": 0,
                }
            _append_channel_preset_event_conn(
                conn,
                assignment_id=int(row["id"]),
                event_type="deleted",
                snapshot=delete_snap,
                details=f"assignment_deleted_by_{reason}",
                metadata={"source": "auto_sync", "reason": reason, **meta},
                ts=now,
            )
            conn.execute("DELETE FROM channel_config_assignments WHERE id = ?", (int(row["id"]),))
            deleted += 1
            continue
        existing_by_pair[key] = row

    real_preset_id = _select_auto_real_preset_id_conn(conn)

    for cid in channel_ids:
        for pid in preset_ids:
            key = (int(cid), int(pid))
            desired_mode = "real" if (real_preset_id is not None and int(pid) == int(real_preset_id)) else "virtual"
            row = existing_by_pair.get(key)
            if row is None:
                conn.execute(
                    """
                    INSERT INTO channel_config_assignments
                    (channel_id, config_id, mode, is_active, created_at, updated_at)
                    VALUES (?, ?, ?, 1, ?, ?)
                    """,
                    (int(cid), int(pid), desired_mode, now, now),
                )
                aid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
                snap = _assignment_snapshot_by_id(conn, aid)
                _append_channel_preset_event_conn(
                    conn,
                    assignment_id=aid,
                    event_type="created",
                    snapshot=snap,
                    details=f"assignment_created_by_{reason}",
                    metadata={"source": "auto_sync", "reason": reason, "mode": desired_mode, **meta},
                    ts=now,
                )
                created += 1
                continue
            prev_mode = str(row["mode"] or "virtual")
            prev_active = int(row["is_active"] or 0)
            if prev_mode != desired_mode:
                conn.execute(
                    """
                    UPDATE channel_config_assignments
                    SET mode = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (desired_mode, now, int(row["id"])),
                )
                snap = _assignment_snapshot_by_id(conn, int(row["id"]))
                _append_channel_preset_event_conn(
                    conn,
                    assignment_id=int(row["id"]),
                    event_type="updated",
                    snapshot=snap,
                    details=f"assignment_updated_by_{reason}",
                    metadata={
                        "source": "auto_sync",
                        "reason": reason,
                        "from_mode": prev_mode,
                        "to_mode": desired_mode,
                        "from_is_active": prev_active,
                        "to_is_active": prev_active,
                        **meta,
                    },
                    ts=now,
                )
                updated += 1

    return {
        "reason": reason,
        "channels": len(channel_ids),
        "presets": len(preset_ids),
        "created": int(created),
        "updated": int(updated),
        "deleted": int(deleted),
        "real_preset_id": int(real_preset_id) if real_preset_id is not None else None,
    }


def _sync_channel_preset_assignments(*, reason: str = "auto_sync", metadata: dict | None = None) -> dict:
    _ensure_channels_table()
    _ensure_operator_presets_table()
    _ensure_assignments_table()
    _ensure_channel_preset_events_table()
    with _db_conn() as conn:
        result = _sync_channel_preset_assignments_conn(conn, reason=reason, metadata=metadata)
        conn.commit()
    return result


def _seed_channel_preset_cross_product() -> dict:
    return _sync_channel_preset_assignments(reason="seed_cross_product")


def _get_operator_defaults():
    presets = _list_operator_presets()
    selected = None
    for p in presets:
        if p.get("is_default"):
            selected = p
            break
    if not selected and presets:
        selected = presets[0]
    if selected:
        return {
            "preset_id": selected["id"],
            "preset_name": selected["name"],
            "mt5_terminal_path": selected["mt5_terminal_path"],
            "mt5_login": selected["mt5_login"],
            "mt5_server": selected["mt5_server"],
            "execution_profile_id": selected["execution_profile_id"],
            "execution_profile_code": selected["execution_profile_code"],
            "execution_profile_name": selected["execution_profile_name"],
            "total_volume": selected["total_volume"],
            "near_entry_pips_min": selected["near_entry_pips_min"],
            "near_entry_spread_mult": selected["near_entry_spread_mult"],
            "verify_order_after_send": selected["verify_order_after_send"],
            "auto_close_on_mismatch": selected["auto_close_on_mismatch"],
        }
    return {
        "preset_id": None,
        "preset_name": "",
        "mt5_terminal_path": MT5_TERMINAL_DEFAULT,
        "mt5_login": None,
        "mt5_server": "",
        "execution_profile_id": None,
        "execution_profile_code": "SWING",
        "execution_profile_name": "Swing",
        "total_volume": 0.03,
        "near_entry_pips_min": 1.0,
        "near_entry_spread_mult": 2.0,
        "verify_order_after_send": True,
        "auto_close_on_mismatch": False,
    }


def _set_default_preset(conn: sqlite3.Connection, preset_id: int) -> None:
    conn.execute("UPDATE operator_presets SET is_default = 0")
    conn.execute("UPDATE operator_presets SET is_default = 1 WHERE id = ?", (preset_id,))


def _validate_single_real_preset_conn(
    conn: sqlite3.Connection,
    *,
    execution_profile_id: int,
    is_real: bool,
    current_preset_id: int | None = None,
) -> None:
    if not bool(is_real):
        return
    prof = conn.execute(
        "SELECT UPPER(COALESCE(code, ?)) AS code FROM execution_profiles WHERE id = ?",
        (DEFAULT_PROFILE_CODE, int(execution_profile_id)),
    ).fetchone()
    if not prof:
        raise HTTPException(status_code=404, detail="Execution profile not found")
    profile_code = str(prof["code"] or DEFAULT_PROFILE_CODE).upper()
    if profile_code != "SWING":
        raise HTTPException(status_code=409, detail="El único preset real debe usar perfil SWING.")
    if current_preset_id is None:
        existing = conn.execute(
            """
            SELECT id, name
            FROM operator_presets
            WHERE is_default = 1
            LIMIT 1
            """
        ).fetchone()
    else:
        existing = conn.execute(
            """
            SELECT id, name
            FROM operator_presets
            WHERE is_default = 1 AND id <> ?
            LIMIT 1
            """,
            (int(current_preset_id),),
        ).fetchone()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=(
                "No se puede marcar otro preset como real. "
                f"Ya existe preset real #{int(existing['id'])} ({str(existing['name'])})."
            ),
        )


def _profile_id_by_code(conn: sqlite3.Connection, code: str) -> int:
    row = conn.execute("SELECT id FROM execution_profiles WHERE UPPER(code) = ?", (str(code).upper(),)).fetchone()
    if row:
        return int(row["id"])
    return _swing_profile_id(conn)


def _seed_recommended_presets() -> None:
    _ensure_operator_presets_table()
    now = _utc_now_iso()
    with _db_conn() as conn:
        base = conn.execute(
            """
            SELECT mt5_terminal_path, mt5_login, mt5_server, total_volume
            FROM operator_presets
            ORDER BY is_default DESC, id ASC
            LIMIT 1
            """
        ).fetchone()
        base_terminal = str(base["mt5_terminal_path"]) if base else str(MT5_TERMINAL_DEFAULT)
        base_login = int(base["mt5_login"]) if base and int(base["mt5_login"] or 0) > 0 else 1
        base_server = str(base["mt5_server"]) if base else "SET_SERVER"
        base_volume = float(base["total_volume"]) if base and float(base["total_volume"] or 0.0) > 0 else 0.01

        recommended = [
            ("REAL_SAFE_FX", "SWING", base_volume, 0.4, 1.2, 1, 0),
            ("REAL_BALANCED_FX", "SWING", max(base_volume, 0.02), 0.8, 1.8, 1, 0),
            ("SCALP_FAST", "SCALP", base_volume, 1.5, 2.5, 1, 0),
            ("SWING_STRICT_PENDING", "SWING", base_volume, 0.2, 1.0, 1, 0),
            ("VIRTUAL_STRESS_2X", "SCALP", base_volume * 2.0, 0.8, 1.8, 1, 0),
            ("VIRTUAL_STRESS_3X", "SCALP", base_volume * 3.0, 0.8, 1.8, 1, 0),
        ]
        for name, profile_code, total_volume, near_min, near_mult, verify_after, auto_close in recommended:
            row = conn.execute("SELECT id FROM operator_presets WHERE name = ?", (name,)).fetchone()
            if row:
                continue
            profile_id = _profile_id_by_code(conn, profile_code)
            conn.execute(
                """
                INSERT INTO operator_presets (
                    name, mt5_terminal_path, mt5_login, mt5_server,
                    execution_profile_id, total_volume, near_entry_pips_min, near_entry_spread_mult,
                    verify_order_after_send, auto_close_on_mismatch, is_default, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
                """,
                (
                    name,
                    base_terminal,
                    base_login,
                    base_server,
                    profile_id,
                    float(total_volume),
                    float(near_min),
                    float(near_mult),
                    int(verify_after),
                    int(auto_close),
                    now,
                    now,
                ),
            )
        conn.commit()


def _migrate_profiles_and_assignments_once() -> None:
    key = "migration_profiles_scalp_swing_v2_done"
    values = _settings_values()
    if _setting_get_bool(values, key, False):
        return
    try:
        _seed_channel_preset_cross_product()
        _setting_set(key, "true")
    except Exception:
        # No bloquear arranque por migración: se podrá reintentar manualmente desde la UI.
        pass


def _list_channels() -> list[dict]:
    _ensure_channels_table()
    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT id, name, chat_id, external_id, is_active, created_at, updated_at
            FROM telegram_channels
            ORDER BY id ASC
            """
        ).fetchall()
    return [
        {
            "id": int(r["id"]),
            "name": str(r["name"]),
            "chat_id": str(r["chat_id"]),
            "external_id": str(r["external_id"] or ""),
            "is_active": bool(r["is_active"]),
            "created_at": str(r["created_at"] or ""),
            "updated_at": str(r["updated_at"] or ""),
        }
        for r in rows
    ]


def _normalize_channel_payload(name: str, chat_id: str, external_id: str, is_active: bool) -> tuple[str, str, str, int]:
    clean_name = (name or "").strip()
    clean_chat_id = (chat_id or "").strip()
    clean_external_id = (external_id or "").strip()
    if not clean_name:
        raise HTTPException(status_code=400, detail="Channel name is required")
    if not clean_chat_id:
        raise HTTPException(status_code=400, detail="Channel chat_id is required")
    try:
        int(clean_chat_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="chat_id must be a valid integer (example: -1001234567890)") from exc
    return clean_name, clean_chat_id, clean_external_id, 1 if is_active else 0


def _active_channels_count() -> int:
    _ensure_channels_table()
    with _db_conn() as conn:
        return int(conn.execute("SELECT COUNT(1) FROM telegram_channels WHERE is_active = 1").fetchone()[0])


def _fetch_report_rows(limit: int = 2000):
    _ensure_reports_tables()
    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                ts, event_id, channel_id, channel_name, config_id, config_name, mode,
                event_type, symbol, side, entry_message_id, reply_to, status,
                error_type, pnl_usd, pnl_pips, details
            FROM strategy_event_log
            ORDER BY ts DESC, id DESC
            LIMIT ?
            """,
            (int(limit),),
        ).fetchall()
    return rows


def _active_assignment_combos() -> dict[str, dict]:
    combos: dict[str, dict] = {}
    for a in _list_assignments():
        if not bool(a.get("is_active")):
            continue
        channel_name = str(a.get("channel_name") or "unknown")
        preset_name = str(a.get("preset_name") or a.get("config_name") or "default")
        key = f"{channel_name}.{preset_name}"
        combos[key] = {
            "combo": key,
            "assignment_id": int(a.get("id") or 0),
            "channel_id": int(a.get("channel_id") or 0),
            "channel_name": channel_name,
            "config_id": int(a.get("config_id") or 0),
            "config_name": preset_name,
            "preset_name": preset_name,
            "mode": str(a.get("mode") or "virtual"),
            "is_active": True,
            "events": 0,
            "entries": 0,
            "modifications": 0,
            "closes": 0,
            "errors": 0,
            "pnl_usd": 0.0,
            "pnl_pips": 0.0,
        }
    return combos


def _channel_preset_registry(assignment_id: int | None = None, from_ts: str | None = None, to_ts: str | None = None):
    _ensure_channel_preset_events_table()
    from_dt = _parse_iso_dt_or_none(from_ts)
    to_dt = _parse_iso_dt_or_none(to_ts)

    where = []
    args: list = []
    if assignment_id is not None:
        where.append("assignment_id = ?")
        args.append(int(assignment_id))
    if from_dt is not None:
        where.append("ts >= ?")
        args.append(from_dt.isoformat(timespec="seconds"))
    if to_dt is not None:
        where.append("ts <= ?")
        args.append(to_dt.isoformat(timespec="seconds"))
    where_sql = (" WHERE " + " AND ".join(where)) if where else ""

    with _db_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT assignment_id, MIN(ts) AS first_seen, MAX(ts) AS last_seen, COUNT(1) AS events
            FROM channel_preset_events
            {where_sql}
            GROUP BY assignment_id
            ORDER BY assignment_id DESC
            """,
            tuple(args),
        ).fetchall()

        items = []
        for r in rows:
            aid = int(r["assignment_id"])
            latest = conn.execute(
                """
                SELECT *
                FROM channel_preset_events
                WHERE assignment_id = ?
                ORDER BY ts DESC, id DESC
                LIMIT 1
                """,
                (aid,),
            ).fetchone()
            if not latest:
                continue
            items.append(
                {
                    "assignment_id": aid,
                    "channel_id": int(latest["channel_id"]) if latest["channel_id"] is not None else None,
                    "channel_name": str(latest["channel_name"] or ""),
                    "channel_chat_id": str(latest["channel_chat_id"] or ""),
                    "config_id": int(latest["config_id"]) if latest["config_id"] is not None else None,
                    "config_name": str(latest["config_name"] or ""),
                    "preset_name": str(latest["config_name"] or ""),
                    "current_mode": str(latest["mode"] or ""),
                    "current_is_active": bool(latest["is_active"]),
                    "current_event_type": str(latest["event_type"] or ""),
                    "first_seen": str(r["first_seen"] or ""),
                    "last_seen": str(r["last_seen"] or ""),
                    "events_count": int(r["events"] or 0),
                }
            )
    return items


def _channel_preset_detail(assignment_id: int, from_ts: str | None = None, to_ts: str | None = None):
    _ensure_channel_preset_events_table()
    _ensure_reports_tables()
    aid = int(assignment_id)
    from_dt = _parse_iso_dt_or_none(from_ts)
    to_dt = _parse_iso_dt_or_none(to_ts)

    with _db_conn() as conn:
        ev_rows = conn.execute(
            """
            SELECT id, ts, event_type, channel_id, channel_name, channel_chat_id, config_id, config_name,
                   mode, is_active, details, metadata_json
            FROM channel_preset_events
            WHERE assignment_id = ?
            ORDER BY ts ASC, id ASC
            """,
            (aid,),
        ).fetchall()

        if not ev_rows:
            current = conn.execute(
                """
                SELECT
                    a.id AS assignment_id, a.created_at AS ts, 'bootstrap' AS event_type,
                    a.channel_id, c.name AS channel_name, c.chat_id AS channel_chat_id,
                    a.config_id, p.name AS config_name, a.mode, a.is_active,
                    'assignment_bootstrap_from_current' AS details,
                    '{}' AS metadata_json
                FROM channel_config_assignments a
                LEFT JOIN telegram_channels c ON c.id = a.channel_id
                LEFT JOIN operator_presets p ON p.id = a.config_id
                WHERE a.id = ?
                """,
                (aid,),
            ).fetchone()
            if current:
                ev_rows = [current]

        if not ev_rows:
            return None

        first_event = ev_rows[0]
        last_event = ev_rows[-1]
        assignment_start = _parse_iso_dt_or_none(str(first_event["ts"] or ""))
        assignment_end = None
        for e in ev_rows:
            if str(e["event_type"] or "").lower() == "deleted":
                assignment_end = _parse_iso_dt_or_none(str(e["ts"] or ""))
        if assignment_start is None:
            assignment_start = datetime.now(URUGUAY_TZ).replace(tzinfo=None)

        channel_id = int(first_event["channel_id"]) if first_event["channel_id"] is not None else None
        config_id = int(first_event["config_id"]) if first_event["config_id"] is not None else None
        current_assignment = _assignment_row_by_id(conn, aid)
        current_assignment_exists = current_assignment is not None
        current_assignment_created_at = str(first_event["ts"] or "")
        current_mode = str(last_event["mode"] or "")
        current_is_active = bool(last_event["is_active"])
        current_open_operations = 0
        if current_assignment is not None:
            current_assignment_created_at = str(current_assignment["created_at"] or first_event["ts"] or "")
            current_mode = str(current_assignment["mode"] or current_mode)
            current_is_active = bool(current_assignment["is_active"])
            if current_assignment["channel_id"] is not None and current_assignment["config_id"] is not None:
                current_open_operations = _open_operations_count_for_channel_preset_conn(
                    conn,
                    int(current_assignment["channel_id"]),
                    int(current_assignment["config_id"]),
                )

        periods = []
        for idx, e in enumerate(ev_rows):
            start_ts = str(e["ts"] or "")
            end_ts = str(ev_rows[idx + 1]["ts"] or "") if idx + 1 < len(ev_rows) else None
            periods.append(
                {
                    "index": idx + 1,
                    "start_ts": start_ts,
                    "end_ts": end_ts,
                    "mode": str(e["mode"] or ""),
                    "is_active": bool(e["is_active"]),
                    "event_type": str(e["event_type"] or ""),
                    "details": str(e["details"] or ""),
                    "metadata_json": str(e["metadata_json"] or "{}"),
                }
            )

        operation_items = []
        modification_items = []
        pnl_series = []
        pips_series = []
        stats = {
            "operations_total": 0,
            "operations_closed": 0,
            "operations_open_pending": 0,
            "modifications_total": 0,
            "pnl_total_usd": 0.0,
            "pnl_total_pips": 0.0,
            "buy_count": 0,
            "sell_count": 0,
            "long_duration_count": 0,
            "avg_duration_sec": 0.0,
            "frequency_trades_per_day": 0.0,
        }

        if channel_id is not None and config_id is not None:
            op_rows = conn.execute(
                """
                SELECT *
                FROM operation_records
                WHERE channel_id = ? AND preset_id = ?
                ORDER BY opened_at ASC, id ASC
                """,
                (channel_id, config_id),
            ).fetchall()

            selected_ops = []
            for r in op_rows:
                opened_raw = str(r["opened_at"] or r["entry_ts"] or "")
                opened_dt = _parse_iso_dt_or_none(opened_raw)
                if opened_dt is None:
                    continue
                if assignment_start is not None and opened_dt < assignment_start:
                    continue
                if assignment_end is not None and opened_dt > assignment_end:
                    continue
                if from_dt is not None and opened_dt < from_dt:
                    continue
                if to_dt is not None and opened_dt > to_dt:
                    continue
                selected_ops.append(r)

            op_ids = []
            duration_sum = 0
            pnl_acc = 0.0
            pips_acc = 0.0
            for r in selected_ops:
                op_ids.append(int(r["id"]))
                duration = int(r["duration_seconds"]) if r["duration_seconds"] is not None else 0
                side = str(r["side"] or "").upper()
                status = str(r["status"] or "")
                pnl_usd = _to_float_or_none(r["pnl_usd"]) or 0.0
                pnl_pips = _to_float_or_none(r["pnl_pips"]) or 0.0
                if side.startswith("B"):
                    stats["buy_count"] += 1
                elif side.startswith("S"):
                    stats["sell_count"] += 1
                if duration >= 4 * 3600:
                    stats["long_duration_count"] += 1
                if status == "CLOSED":
                    stats["operations_closed"] += 1
                else:
                    stats["operations_open_pending"] += 1
                stats["operations_total"] += 1
                stats["modifications_total"] += int(r["modifications_count"] or 0)
                stats["pnl_total_usd"] += float(pnl_usd)
                stats["pnl_total_pips"] += float(pnl_pips)
                duration_sum += int(max(0, duration))

                close_ref = str(r["closed_at"] or r["updated_at"] or r["opened_at"] or "")
                pnl_acc += float(pnl_usd)
                pips_acc += float(pnl_pips)
                pnl_series.append({"ts": close_ref, "value": float(pnl_acc)})
                pips_series.append({"ts": close_ref, "value": float(pips_acc)})

                operation_items.append(
                    {
                        "id": int(r["id"]),
                        "status": status,
                        "mode": str(r["mode"] or ""),
                        "is_virtual": bool(r["is_virtual"]),
                        "symbol": str(r["symbol"] or ""),
                        "side": str(r["side"] or ""),
                        "opened_at": str(r["opened_at"] or ""),
                        "closed_at": str(r["closed_at"] or ""),
                        "entry_message_id": str(r["entry_message_id"] or ""),
                        "entry_event_id": str(r["entry_event_id"] or ""),
                        "modifications_count": int(r["modifications_count"] or 0),
                        "last_modification_message_id": str(r["last_modification_message_id"] or ""),
                        "last_modified_at": str(r["last_modified_at"] or ""),
                        "duration_seconds": int(r["duration_seconds"]) if r["duration_seconds"] is not None else None,
                        "pnl_usd": _to_float_or_none(r["pnl_usd"]),
                        "pnl_pips": _to_float_or_none(r["pnl_pips"]),
                    }
                )

            if op_ids:
                placeholders = ",".join(["?"] * len(op_ids))
                mod_rows = conn.execute(
                    f"""
                    SELECT operation_id, ts, event_type, event_id, message_id, status, error_type, sl, tp, pnl_usd, pnl_pips, details
                    FROM operation_events
                    WHERE operation_id IN ({placeholders})
                      AND event_type = 'modification'
                    ORDER BY ts ASC, id ASC
                    """,
                    tuple(op_ids),
                ).fetchall()
                for m in mod_rows:
                    mts = str(m["ts"] or "")
                    mts_dt = _parse_iso_dt_or_none(mts)
                    if from_dt is not None and mts_dt is not None and mts_dt < from_dt:
                        continue
                    if to_dt is not None and mts_dt is not None and mts_dt > to_dt:
                        continue
                    modification_items.append(
                        {
                            "operation_id": int(m["operation_id"]),
                            "ts": mts,
                            "event_id": str(m["event_id"] or ""),
                            "message_id": str(m["message_id"] or ""),
                            "status": str(m["status"] or ""),
                            "error_type": str(m["error_type"] or ""),
                            "sl": _to_float_or_none(m["sl"]),
                            "tp": _to_float_or_none(m["tp"]),
                            "pnl_usd": _to_float_or_none(m["pnl_usd"]),
                            "pnl_pips": _to_float_or_none(m["pnl_pips"]),
                            "details": str(m["details"] or ""),
                        }
                    )

            if stats["operations_total"] > 0:
                stats["avg_duration_sec"] = float(duration_sum) / float(stats["operations_total"])

        now_dt = datetime.now(URUGUAY_TZ).replace(tzinfo=None)
        end_for_freq = assignment_end if assignment_end is not None else now_dt
        active_days = max(1.0, (end_for_freq - assignment_start).total_seconds() / 86400.0)
        stats["frequency_trades_per_day"] = float(stats["operations_total"]) / active_days

    return {
        "assignment_id": aid,
        "meta": {
            "channel_id": channel_id,
            "channel_name": str(first_event["channel_name"] or ""),
            "channel_chat_id": str(first_event["channel_chat_id"] or ""),
            "config_id": config_id,
            "config_name": str(first_event["config_name"] or ""),
            "preset_name": str(first_event["config_name"] or ""),
            "created_at": current_assignment_created_at,
            "first_seen": str(first_event["ts"] or ""),
            "last_seen": str(last_event["ts"] or ""),
            "deleted_at": assignment_end.isoformat(timespec="seconds") if assignment_end is not None else "",
            "current_mode": current_mode,
            "current_is_active": current_is_active,
            "current_event_type": str(last_event["event_type"] or ""),
            "current_assignment_exists": bool(current_assignment_exists),
            "current_open_operations": int(current_open_operations),
        },
        "periods": periods,
        "events": [
            {
                "id": int(e["id"]) if "id" in e.keys() else None,
                "ts": str(e["ts"] or ""),
                "event_type": str(e["event_type"] or ""),
                "mode": str(e["mode"] or ""),
                "is_active": bool(e["is_active"]),
                "details": str(e["details"] or ""),
                "metadata_json": str(e["metadata_json"] or "{}"),
            }
            for e in ev_rows
        ],
        "operations": operation_items,
        "modifications": modification_items,
        "pnl_series": pnl_series,
        "pips_series": pips_series,
        "stats": stats,
    }


def _channel_preset_metrics_for_active() -> list[dict]:
    _ensure_reports_tables()
    seeds = _active_assignment_combos()
    if not seeds:
        return []
    by_combo = {k: dict(v) for k, v in seeds.items()}
    for item in by_combo.values():
        item["avg_duration_sec"] = 0.0
        item["long_duration_count"] = 0
        item["buy_count"] = 0
        item["sell_count"] = 0
        item["frequency_trades_per_day"] = 0.0

    with _db_conn() as conn:
        for item in by_combo.values():
            rows = conn.execute(
                """
                SELECT side, status, opened_at, closed_at, duration_seconds, pnl_usd, pnl_pips
                FROM operation_records
                WHERE channel_id = ? AND preset_id = ?
                ORDER BY opened_at ASC, id ASC
                """,
                (int(item["channel_id"]), int(item["config_id"])),
            ).fetchall()
            if not rows:
                continue
            total = 0
            dur_sum = 0
            long_count = 0
            first_dt = None
            last_dt = None
            for r in rows:
                total += 1
                side = str(r["side"] or "").upper()
                if side.startswith("B"):
                    item["buy_count"] += 1
                elif side.startswith("S"):
                    item["sell_count"] += 1
                dur = int(r["duration_seconds"]) if r["duration_seconds"] is not None else 0
                dur_sum += max(0, dur)
                if dur >= 4 * 3600:
                    long_count += 1
                dt = _parse_iso_dt_or_none(str(r["opened_at"] or ""))
                if dt is not None:
                    if first_dt is None or dt < first_dt:
                        first_dt = dt
                    if last_dt is None or dt > last_dt:
                        last_dt = dt
            item["entries"] = total
            item["long_duration_count"] = long_count
            item["avg_duration_sec"] = (float(dur_sum) / float(total)) if total > 0 else 0.0
            if first_dt is not None and last_dt is not None:
                days = max(1.0, (last_dt - first_dt).total_seconds() / 86400.0)
                item["frequency_trades_per_day"] = float(total) / days
    return list(by_combo.values())


def _to_float_or_none(value):
    try:
        if value is None:
            return None
        s = str(value).strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None


def _elapsed_seconds(start_ts: str | None, end_ts: str | None = None):
    start_dt = _parse_iso_dt_or_none(start_ts)
    if start_dt is None:
        return None
    end_dt = _parse_iso_dt_or_none(end_ts) if end_ts else datetime.now(URUGUAY_TZ).replace(tzinfo=None)
    if end_dt is None:
        end_dt = datetime.now(URUGUAY_TZ).replace(tzinfo=None)
    try:
        return max(0, int((end_dt - start_dt).total_seconds()))
    except Exception:
        return None


def _operation_row_to_dict(row: sqlite3.Row, now_ts: str):
    opened_at = str(row["opened_at"] or "")
    closed_at = str(row["closed_at"] or "")
    elapsed = _elapsed_seconds(opened_at, closed_at or now_ts)
    return {
        "id": int(row["id"]),
        "operation_key": str(row["operation_key"] or ""),
        "mode": str(row["mode"] or ""),
        "status": str(row["status"] or ""),
        "is_virtual": bool(row["is_virtual"]),
        "channel_id": int(row["channel_id"]) if row["channel_id"] is not None else None,
        "channel_name": str(row["channel_name"] or ""),
        "channel_index": str(row["channel_index"] or ""),
        "preset_id": int(row["preset_id"]) if row["preset_id"] is not None else None,
        "preset_name": str(row["preset_name"] or ""),
        "execution_profile": str(row["execution_profile"] or ""),
        "symbol": str(row["symbol"] or ""),
        "side": str(row["side"] or ""),
        "entry_message_id": str(row["entry_message_id"] or ""),
        "entry_event_id": str(row["entry_event_id"] or ""),
        "entry_trigger_message_id": str(row["entry_trigger_message_id"] or ""),
        "entry_ts": str(row["entry_ts"] or ""),
        "opened_at": opened_at,
        "ticket": str(row["ticket"] or ""),
        "comment": str(row["comment"] or ""),
        "volume": _to_float_or_none(row["volume"]),
        "entry_price": _to_float_or_none(row["entry_price"]),
        "sl": _to_float_or_none(row["sl"]),
        "tp": _to_float_or_none(row["tp"]),
        "had_modifications": bool(row["had_modifications"]),
        "modifications_count": int(row["modifications_count"] or 0),
        "last_modification_message_id": str(row["last_modification_message_id"] or ""),
        "last_modified_at": str(row["last_modified_at"] or ""),
        "closed_at": closed_at,
        "close_event_id": str(row["close_event_id"] or ""),
        "close_trigger_message_id": str(row["close_trigger_message_id"] or ""),
        "close_reason": str(row["close_reason"] or ""),
        "close_source": str(row["close_source"] or ""),
        "close_error_id": str(row["close_error_id"] or ""),
        "close_error_type": str(row["close_error_type"] or ""),
        "close_details": str(row["close_details"] or ""),
        "pnl_usd": _to_float_or_none(row["pnl_usd"]),
        "pnl_pips": _to_float_or_none(row["pnl_pips"]),
        "last_pips": _to_float_or_none(row["last_pips"]),
        "last_profit_usd": _to_float_or_none(row["last_profit_usd"]),
        "duration_seconds": int(row["duration_seconds"]) if row["duration_seconds"] is not None else elapsed,
        "last_sync_at": str(row["last_sync_at"] or ""),
        "created_at": str(row["created_at"] or ""),
        "updated_at": str(row["updated_at"] or ""),
        "elapsed_seconds": elapsed,
    }


def _fetch_open_operations():
    _ensure_reports_tables()
    now_ts = _utc_now_iso()
    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM operation_records
            WHERE status IN ('OPEN', 'PENDING')
            ORDER BY
                CASE status WHEN 'OPEN' THEN 0 ELSE 1 END,
                opened_at DESC,
                id DESC
            """
        ).fetchall()
    return [_operation_row_to_dict(r, now_ts) for r in rows]


def _fetch_closed_operations(
    page: int,
    page_size: int,
    *,
    opened_from_ts: str | None = None,
    opened_to_ts: str | None = None,
    closed_from_ts: str | None = None,
    closed_to_ts: str | None = None,
    symbol: str | None = None,
    side: str | None = None,
    operation_id: int | None = None,
    channel_preset: str | None = None,
    close_source: str | None = None,
    close_message_id: str | None = None,
    error_id: str | None = None,
):
    _ensure_reports_tables()
    p = max(1, int(page))
    ps = max(1, min(int(page_size), 200))
    where = ["status = 'CLOSED'"]
    args: list = []

    opened_from_dt = _parse_iso_dt_or_none(opened_from_ts)
    opened_to_dt = _parse_iso_dt_or_none(opened_to_ts)
    closed_from_dt = _parse_iso_dt_or_none(closed_from_ts)
    closed_to_dt = _parse_iso_dt_or_none(closed_to_ts)
    symbol_q = str(symbol or "").strip().upper()
    side_q = str(side or "").strip().upper()
    channel_preset_q = str(channel_preset or "").strip()
    close_source_q = str(close_source or "").strip()
    close_message_id_q = str(close_message_id or "").strip()
    error_id_q = str(error_id or "").strip()

    if operation_id is not None:
        where.append("id = ?")
        args.append(int(operation_id))
    if opened_from_dt is not None:
        where.append("COALESCE(opened_at, entry_ts, '') >= ?")
        args.append(opened_from_dt.isoformat(timespec="seconds"))
    if opened_to_dt is not None:
        where.append("COALESCE(opened_at, entry_ts, '') <= ?")
        args.append(opened_to_dt.isoformat(timespec="seconds"))
    if closed_from_dt is not None:
        where.append("COALESCE(closed_at, updated_at, '') >= ?")
        args.append(closed_from_dt.isoformat(timespec="seconds"))
    if closed_to_dt is not None:
        where.append("COALESCE(closed_at, updated_at, '') <= ?")
        args.append(closed_to_dt.isoformat(timespec="seconds"))
    if symbol_q:
        where.append("UPPER(COALESCE(symbol, '')) LIKE ?")
        args.append(f"%{symbol_q}%")
    if side_q in {"BUY", "SELL"}:
        where.append("UPPER(COALESCE(side, '')) = ?")
        args.append(side_q)
    if channel_preset_q:
        where.append("(COALESCE(channel_name, '') || '.' || COALESCE(preset_name, '')) LIKE ?")
        args.append(f"%{channel_preset_q}%")
    if close_source_q:
        where.append("COALESCE(close_source, '') LIKE ?")
        args.append(f"%{close_source_q}%")
    if close_message_id_q:
        where.append("COALESCE(close_trigger_message_id, '') LIKE ?")
        args.append(f"%{close_message_id_q}%")
    if error_id_q:
        where.append("COALESCE(close_error_id, '') LIKE ?")
        args.append(f"%{error_id_q}%")

    where_sql = " AND ".join(where)
    with _db_conn() as conn:
        total = int(
            conn.execute(
                f"SELECT COUNT(1) FROM operation_records WHERE {where_sql}",
                tuple(args),
            ).fetchone()[0]
        )
        total_pages = max(1, (total + ps - 1) // ps)
        if p > total_pages:
            p = total_pages
        start = (p - 1) * ps
        rows = conn.execute(
            f"""
            SELECT *
            FROM operation_records
            WHERE {where_sql}
            ORDER BY closed_at DESC, id DESC
            LIMIT ? OFFSET ?
            """,
            tuple(args + [ps, start]),
        ).fetchall()
    now_ts = _utc_now_iso()
    return {
        "items": [_operation_row_to_dict(r, now_ts) for r in rows],
        "page": p,
        "page_size": ps,
        "total": total,
        "total_pages": total_pages,
        "has_prev": p > 1,
        "has_next": p < total_pages,
        "filters": {
            "opened_from_ts": opened_from_dt.isoformat(timespec="seconds") if opened_from_dt is not None else "",
            "opened_to_ts": opened_to_dt.isoformat(timespec="seconds") if opened_to_dt is not None else "",
            "closed_from_ts": closed_from_dt.isoformat(timespec="seconds") if closed_from_dt is not None else "",
            "closed_to_ts": closed_to_dt.isoformat(timespec="seconds") if closed_to_dt is not None else "",
            "symbol": symbol_q,
            "side": side_q if side_q in {"BUY", "SELL"} else "",
            "operation_id": int(operation_id) if operation_id is not None else None,
            "channel_preset": channel_preset_q,
            "close_source": close_source_q,
            "close_message_id": close_message_id_q,
            "error_id": error_id_q,
        },
    }


def _fetch_operation_detail(operation_id: int):
    _ensure_reports_tables()
    with _db_conn() as conn:
        row = conn.execute("SELECT * FROM operation_records WHERE id = ?", (int(operation_id),)).fetchone()
        if not row:
            return None
        events = conn.execute(
            """
            SELECT id, ts, event_type, event_id, message_id, reply_to, status, error_type,
                   sl, tp, pnl_usd, pnl_pips, details
            FROM operation_events
            WHERE operation_id = ?
            ORDER BY ts ASC, id ASC
            """,
            (int(operation_id),),
        ).fetchall()
    now_ts = _utc_now_iso()
    return {
        "operation": _operation_row_to_dict(row, now_ts),
        "events": [
            {
                "id": int(e["id"]),
                "ts": str(e["ts"] or ""),
                "event_type": str(e["event_type"] or ""),
                "event_id": str(e["event_id"] or ""),
                "message_id": str(e["message_id"] or ""),
                "reply_to": str(e["reply_to"] or ""),
                "status": str(e["status"] or ""),
                "error_type": str(e["error_type"] or ""),
                "sl": _to_float_or_none(e["sl"]),
                "tp": _to_float_or_none(e["tp"]),
                "pnl_usd": _to_float_or_none(e["pnl_usd"]),
                "pnl_pips": _to_float_or_none(e["pnl_pips"]),
                "details": str(e["details"] or ""),
            }
            for e in events
        ],
    }


def _normalize_manual_close_mode(value: str | None) -> str:
    mode = str(value or "all").strip().lower()
    if mode not in {"all", "real", "virtual"}:
        raise HTTPException(status_code=400, detail="mode inválido: all|real|virtual")
    return mode


def _operation_close_defaults(reason: str | None, details: str | None) -> tuple[str, str]:
    close_reason = str(reason or "").strip() or "Cerrada desde Panel web a mano"
    close_details = str(details or "").strip()
    return close_reason, close_details


def _insert_strategy_close_log(conn: sqlite3.Connection, row: sqlite3.Row, *, ts: str, event_id: str, details: str) -> None:
    conn.execute(
        """
        INSERT INTO strategy_event_log (
            ts, event_id, message_id, channel_id, channel_name, config_id, config_name, mode,
            event_type, symbol, side, operator_class, entry_message_id, reply_to, status, error_type,
            pnl_usd, pnl_pips, details
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            ts,
            str(event_id),
            "",
            int(row["channel_id"]) if row["channel_id"] is not None else None,
            str(row["channel_name"] or ""),
            int(row["preset_id"]) if row["preset_id"] is not None else None,
            str(row["preset_name"] or ""),
            str(row["mode"] or ""),
            "close",
            str(row["symbol"] or ""),
            str(row["side"] or ""),
            str(row["execution_profile"] or ""),
            str(row["entry_message_id"] or ""),
            "",
            "CLOSED",
            "",
            _to_float_or_none(row["pnl_usd"]),
            _to_float_or_none(row["pnl_pips"]),
            str(details or ""),
        ),
    )


def _enqueue_panel_close_event(operation_row: sqlite3.Row, *, reason: str, details: str) -> str:
    QUEUE_PENDING_DIR.mkdir(parents=True, exist_ok=True)
    op_id = int(operation_row["id"])
    event_id = f"panel_close_{op_id}_{uuid.uuid4().hex}"
    now_ts = _utc_now_iso()
    payload = {
        "event_id": event_id,
        "type": "panel_close",
        "timestamp": now_ts,
        "message_id": f"panel-close-{op_id}-{int(time.time() * 1000)}",
        "reply_to": str(operation_row["entry_message_id"] or ""),
        "channel": str(operation_row["channel_name"] or ""),
        "channel_index": str(operation_row["channel_index"] or ""),
        "channel_id": int(operation_row["channel_id"]) if operation_row["channel_id"] is not None else "",
        "symbol": str(operation_row["symbol"] or ""),
        "operation": str(operation_row["side"] or ""),
        "operator_class": str(operation_row["execution_profile"] or ""),
        "operation_id": int(op_id),
        "close_reason": str(reason or ""),
        "close_details": str(details or ""),
        "message_text": f"panel_close operation_id={op_id}",
    }
    tmp_path = QUEUE_PENDING_DIR / f"{event_id}.json.tmp"
    final_path = QUEUE_PENDING_DIR / f"{event_id}.json"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=True, default=str)
    os.replace(tmp_path, final_path)
    return event_id


def _close_operation_row_manual_conn(
    conn: sqlite3.Connection,
    row: sqlite3.Row,
    *,
    close_reason: str,
    close_details: str,
) -> int:
    op_id = int(row["id"])
    now = _utc_now_iso()
    close_event_id = f"panel_manual_close:{op_id}:{int(time.time() * 1000)}"
    final_details = close_details or "Cierre manual solicitado desde Operaciones (Panel web)."

    pnl_usd = _to_float_or_none(row["pnl_usd"])
    if pnl_usd is None:
        pnl_usd = _to_float_or_none(row["last_profit_usd"])
    pnl_pips = _to_float_or_none(row["pnl_pips"])
    if pnl_pips is None:
        pnl_pips = _to_float_or_none(row["last_pips"])

    duration = _elapsed_seconds(str(row["opened_at"] or ""), now)
    if duration is None:
        duration = int(row["duration_seconds"]) if row["duration_seconds"] is not None else None

    conn.execute(
        """
        UPDATE operation_records
        SET status = 'CLOSED',
            closed_at = ?,
            close_event_id = ?,
            close_trigger_message_id = '',
            close_reason = ?,
            close_source = 'panel_web_manual',
            close_error_id = '',
            close_error_type = '',
            close_details = ?,
            pnl_usd = ?,
            pnl_pips = ?,
            duration_seconds = ?,
            last_sync_at = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (
            now,
            close_event_id,
            str(close_reason),
            str(final_details),
            pnl_usd,
            pnl_pips,
            duration,
            now,
            now,
            op_id,
        ),
    )
    if str(row["mode"] or "").lower() == "virtual":
        conn.execute(
            """
            UPDATE virtual_positions
            SET status = 'CLOSED',
                closed_at = COALESCE(closed_at, ?),
                close_price = COALESCE(close_price, ?),
                pnl_usd = COALESCE(pnl_usd, ?),
                pnl_pips = COALESCE(pnl_pips, ?)
            WHERE channel_id = ?
              AND config_id = ?
              AND entry_message_id = ?
              AND status = 'OPEN'
            """,
            (
                now,
                _to_float_or_none(row["entry_price"]),
                pnl_usd,
                pnl_pips,
                int(row["channel_id"]) if row["channel_id"] is not None else -1,
                int(row["preset_id"]) if row["preset_id"] is not None else -1,
                str(row["entry_message_id"] or ""),
            ),
        )

    conn.execute(
        """
        INSERT INTO operation_events (
            operation_id, ts, event_type, event_id, message_id, reply_to, status, error_type,
            sl, tp, pnl_usd, pnl_pips, details
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            op_id,
            now,
            "close",
            close_event_id,
            "",
            "",
            "CLOSED",
            "",
            _to_float_or_none(row["sl"]),
            _to_float_or_none(row["tp"]),
            pnl_usd,
            pnl_pips,
            str(final_details),
        ),
    )

    refreshed = conn.execute("SELECT * FROM operation_records WHERE id = ?", (op_id,)).fetchone()
    if refreshed:
        _insert_strategy_close_log(conn, refreshed, ts=now, event_id=close_event_id, details=final_details)
    return op_id


def _close_operation_manual(
    operation_id: int,
    *,
    reason: str | None = None,
    details: str | None = None,
    close_in_mt5: bool = False,
) -> dict:
    _ensure_reports_tables()
    close_reason, close_details = _operation_close_defaults(reason, details)
    with _db_conn() as conn:
        row = conn.execute("SELECT * FROM operation_records WHERE id = ?", (int(operation_id),)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Operation not found")
        status = str(row["status"] or "").upper()
        if status not in {"OPEN", "PENDING"}:
            raise HTTPException(status_code=409, detail=f"La operación #{int(operation_id)} no está abierta/pending (status={status})")
        mode = str(row["mode"] or "").lower()
        if bool(close_in_mt5) and mode == "real":
            if not operador_manager.running():
                raise HTTPException(status_code=409, detail="Operador no está corriendo. Inícialo para cierre real en MT5.")
            queued_event_id = _enqueue_panel_close_event(
                row,
                reason=close_reason,
                details=close_details or "Cierre solicitado desde panel para ejecutar en MT5.",
            )
            conn.execute(
                """
                INSERT INTO operation_events (
                    operation_id, ts, event_type, event_id, message_id, reply_to, status, error_type,
                    sl, tp, pnl_usd, pnl_pips, details
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(operation_id),
                    _utc_now_iso(),
                    "close",
                    queued_event_id,
                    "",
                    "",
                    "QUEUED",
                    "",
                    _to_float_or_none(row["sl"]),
                    _to_float_or_none(row["tp"]),
                    None,
                    None,
                    "panel_manual_mt5_queued",
                ),
            )
            conn.commit()
            now_ts = _utc_now_iso()
            refreshed = conn.execute("SELECT * FROM operation_records WHERE id = ?", (int(operation_id),)).fetchone()
            return {
                "closed_count": 0,
                "closed_ids": [],
                "queued_count": 1,
                "queued_event_ids": [queued_event_id],
                "operation": _operation_row_to_dict(refreshed, now_ts) if refreshed else None,
                "close_source": "panel_web_manual_mt5_queued",
                "close_reason": close_reason,
            }
        op_id = _close_operation_row_manual_conn(
            conn,
            row,
            close_reason=close_reason,
            close_details=close_details,
        )
        conn.commit()
        updated = conn.execute("SELECT * FROM operation_records WHERE id = ?", (int(op_id),)).fetchone()
    now_ts = _utc_now_iso()
    return {
        "closed_count": 1,
        "closed_ids": [int(op_id)],
        "operation": _operation_row_to_dict(updated, now_ts) if updated else None,
        "close_source": "panel_web_manual",
        "close_reason": close_reason,
    }


def _close_operations_manual_bulk(
    *,
    mode: str = "all",
    include_pending: bool = True,
    reason: str | None = None,
    details: str | None = None,
    close_in_mt5: bool = False,
) -> dict:
    _ensure_reports_tables()
    mode_filter = _normalize_manual_close_mode(mode)
    close_reason, close_details = _operation_close_defaults(reason, details)
    statuses = ["OPEN", "PENDING"] if bool(include_pending) else ["OPEN"]

    where = [f"status IN ({','.join(['?'] * len(statuses))})"]
    params: list = list(statuses)
    if mode_filter in {"real", "virtual"}:
        where.append("mode = ?")
        params.append(mode_filter)

    with _db_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT *
            FROM operation_records
            WHERE {' AND '.join(where)}
            ORDER BY id ASC
            """,
            tuple(params),
        ).fetchall()

        closed_ids: list[int] = []
        queued_ids: list[str] = []
        if bool(close_in_mt5):
            has_real = any(str(r["mode"] or "").lower() == "real" for r in rows)
            if has_real and not operador_manager.running():
                raise HTTPException(status_code=409, detail="Operador no está corriendo. Inícialo para cierre real en MT5.")
        for row in rows:
            try:
                if bool(close_in_mt5) and str(row["mode"] or "").lower() == "real":
                    event_id = _enqueue_panel_close_event(
                        row,
                        reason=close_reason,
                        details=close_details or "Cierre masivo solicitado desde panel para ejecutar en MT5.",
                    )
                    queued_ids.append(str(event_id))
                    conn.execute(
                        """
                        INSERT INTO operation_events (
                            operation_id, ts, event_type, event_id, message_id, reply_to, status, error_type,
                            sl, tp, pnl_usd, pnl_pips, details
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            int(row["id"]),
                            _utc_now_iso(),
                            "close",
                            str(event_id),
                            "",
                            "",
                            "QUEUED",
                            "",
                            _to_float_or_none(row["sl"]),
                            _to_float_or_none(row["tp"]),
                            None,
                            None,
                            "panel_manual_mt5_queued_bulk",
                        ),
                    )
                    continue
                op_id = _close_operation_row_manual_conn(
                    conn,
                    row,
                    close_reason=close_reason,
                    close_details=close_details,
                )
                closed_ids.append(int(op_id))
            except Exception:
                continue
        conn.commit()

    return {
        "mode": mode_filter,
        "include_pending": bool(include_pending),
        "matched": len(rows),
        "closed_count": len(closed_ids),
        "closed_ids": closed_ids,
        "queued_count": len(queued_ids),
        "queued_event_ids": queued_ids,
        "close_source": "panel_web_manual_mt5_queued" if bool(close_in_mt5) else "panel_web_manual",
        "close_reason": close_reason,
    }


def _search_messages(query: str, limit: int):
    _ensure_reports_tables()
    q = str(query or "").strip()
    lim = max(1, min(int(limit), 500))
    if not q:
        return []
    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                id, message_uid, message_key, message_id, channel_name, channel_index, ts, reply_to,
                event_id, event_type, symbol, operation, operator_class, message_text, raw_payload
            FROM telegram_messages
            WHERE message_id = ?
               OR event_id = ?
               OR message_uid = ?
               OR message_key = ?
               OR message_uid LIKE ?
            ORDER BY ts DESC, id DESC
            LIMIT ?
            """,
            (q, q, q, q, f"{q}:%", lim),
        ).fetchall()
        extra = conn.execute(
            """
            SELECT
                id, ts, event_id, message_id, channel_name, event_type, symbol, side, operator_class, reply_to, details
            FROM strategy_event_log
            WHERE event_id = ?
               OR message_id = ?
            ORDER BY ts DESC, id DESC
            LIMIT ?
            """,
            (q, q, lim),
        ).fetchall()
    items = [
        {
            "id": int(r["id"]),
            "message_uid": str(r["message_uid"] or ""),
            "message_key": str(r["message_key"] or ""),
            "message_id": str(r["message_id"] or ""),
            "channel_name": str(r["channel_name"] or ""),
            "channel_index": str(r["channel_index"] or ""),
            "ts": str(r["ts"] or ""),
            "reply_to": str(r["reply_to"] or ""),
            "event_id": str(r["event_id"] or ""),
            "event_type": str(r["event_type"] or ""),
            "symbol": str(r["symbol"] or ""),
            "operation": str(r["operation"] or ""),
            "operator_class": str(r["operator_class"] or ""),
            "message_text": str(r["message_text"] or ""),
            "raw_payload": str(r["raw_payload"] or ""),
        }
        for r in rows
    ]
    seen = {(str(x.get("message_key") or x["message_uid"]), str(x["event_id"])) for x in items}
    for r in extra:
        msg_id = str(r["message_id"] or "")
        event_id = str(r["event_id"] or "")
        chan = str(r["channel_name"] or "")
        message_key = f"{chan}:{msg_id}" if msg_id else (f"event:{event_id}" if event_id else "")
        message_uid = message_key
        key = (message_key, event_id)
        if key in seen:
            continue
        seen.add(key)
        items.append(
            {
                "id": int(r["id"]),
                "message_uid": message_uid,
                "message_key": message_key,
                "message_id": msg_id,
                "channel_name": chan,
                "channel_index": "",
                "ts": str(r["ts"] or ""),
                "reply_to": str(r["reply_to"] or ""),
                "event_id": event_id,
                "event_type": str(r["event_type"] or ""),
                "symbol": str(r["symbol"] or ""),
                "operation": str(r["side"] or ""),
                "operator_class": str(r["operator_class"] or ""),
                "message_text": "",
                "raw_payload": str(r["details"] or ""),
            }
        )
    items.sort(key=lambda x: str(x.get("ts", "")), reverse=True)
    return items[:lim]


restart_lock = threading.Lock()
alerts_lock = threading.Lock()
workers_started = False


def _restart_settings_snapshot() -> dict:
    raw = _settings_values()
    target = str(raw.get("auto_restart_target", "operador")).strip().lower()
    if target not in RESTART_TARGETS:
        target = "operador"
    enabled = _setting_get_bool(raw, "auto_restart_enabled", False)
    interval_min = max(5, _setting_get_int(raw, "auto_restart_interval_min", 240))
    next_at = str(raw.get("auto_restart_next_at", "") or "").strip()
    next_dt = _parse_iso_utc_or_none(next_at)
    if enabled and next_dt is None:
        next_at = _utc_now_iso()
        next_dt = _parse_iso_utc_or_none(next_at)
        _setting_set("auto_restart_next_at", next_at)
    now = datetime.now(URUGUAY_TZ).replace(tzinfo=None)
    remaining = int((next_dt - now).total_seconds()) if next_dt else None
    return {
        "enabled": enabled,
        "interval_minutes": interval_min,
        "target": target,
        "next_restart_at": next_at,
        "seconds_remaining": max(0, remaining) if remaining is not None else None,
    }


def _alerts_settings_snapshot() -> dict:
    raw = _settings_values()
    min_severity = str(raw.get("discord_min_severity", "warning")).strip().lower()
    if min_severity not in ALERT_SEVERITY_RANK:
        min_severity = "warning"
    return {
        "alerts_enabled": _setting_get_bool(raw, "alerts_enabled", True),
        "alerts_check_interval_sec": max(5, _setting_get_int(raw, "alerts_check_interval_sec", 10)),
        "alerts_queue_pending_threshold": max(1, _setting_get_int(raw, "alerts_queue_pending_threshold", 50)),
        "alerts_queue_oldest_sec": max(30, _setting_get_int(raw, "alerts_queue_oldest_sec", 180)),
        "alerts_pending_order_sec": max(60, _setting_get_int(raw, "alerts_pending_order_sec", 1200)),
        "alerts_error_window_min": max(1, _setting_get_int(raw, "alerts_error_window_min", 15)),
        "alerts_error_count_threshold": max(1, _setting_get_int(raw, "alerts_error_count_threshold", 8)),
        "alerts_no_tickets_threshold": max(1, _setting_get_int(raw, "alerts_no_tickets_threshold", 3)),
        "alerts_drawdown_daily_usd": _setting_get_float(raw, "alerts_drawdown_daily_usd", -150.0),
        "alerts_stale_sync_sec": max(30, _setting_get_int(raw, "alerts_stale_sync_sec", 60)),
        "discord_enabled": _setting_get_bool(raw, "discord_enabled", False),
        "discord_webhook_url": str(raw.get("discord_webhook_url", "") or "").strip(),
        "discord_min_severity": min_severity,
    }


def _set_restart_config(enabled: bool, interval_minutes: int, target: str) -> dict:
    t = str(target or "operador").strip().lower()
    if t not in RESTART_TARGETS:
        raise HTTPException(status_code=400, detail="target inválido: operador|lector|both")
    interval = max(5, int(interval_minutes))
    _setting_set("auto_restart_enabled", "true" if bool(enabled) else "false")
    _setting_set("auto_restart_interval_min", str(interval))
    _setting_set("auto_restart_target", t)
    next_at = ""
    if bool(enabled):
        next_at = _uy_from_epoch_iso(time.time() + (interval * 60))
    _setting_set("auto_restart_next_at", str(next_at))
    return _restart_settings_snapshot()


def _restart_due_in_seconds() -> int | None:
    snap = _restart_settings_snapshot()
    if not snap["enabled"]:
        return None
    return snap["seconds_remaining"]


def _perform_controlled_restart(target: str, reason: str = "manual") -> dict:
    t = str(target or "operador").strip().lower()
    if t not in RESTART_TARGETS:
        raise HTTPException(status_code=400, detail="target inválido: operador|lector|both")
    with restart_lock:
        queue_stats = _queue_pending_stats()
        # Regla conservadora: no reiniciar programáticamente cuando hay backlog.
        if reason == "scheduled" and int(queue_stats["count"]) > 0:
            postpone_dt = _uy_from_epoch_iso(time.time() + 60)
            _setting_set("auto_restart_next_at", postpone_dt)
            _upsert_alert(
                source="scheduler",
                code="RESTART_POSTPONED_QUEUE",
                severity="warning",
                title="Reinicio postergado por cola pendiente",
                details=f"Hay {queue_stats['count']} evento(s) en queue/pending. Se reintenta en 60s.",
                data={"target": t, "reason": reason, "queue": queue_stats},
            )
            return {"status": "postponed", "target": t, "reason": reason, "queue": queue_stats, "next_restart_at": postpone_dt}

        actions: list[str] = []
        errors: list[str] = []
        if t in ("operador", "both"):
            try:
                if not operador_manager.last_env:
                    operador_manager.last_env = _restore_process_runtime_env(PROCESS_NAME_OPERADOR)
                if not operador_manager.last_env:
                    raise RuntimeError("OPERADOR sin configuración previa para reiniciar")
                operador_manager.restart()
                actions.append("operador_restarted")
            except Exception as exc:
                errors.append(f"operador: {exc}")
        if t in ("lector", "both"):
            try:
                if not lector_manager.last_env:
                    lector_manager.last_env = _restore_process_runtime_env(PROCESS_NAME_LECTOR)
                if not lector_manager.last_env:
                    raise RuntimeError("LECTOR sin configuración previa para reiniciar")
                lector_manager.restart()
                actions.append("lector_restarted")
            except Exception as exc:
                errors.append(f"lector: {exc}")

        status = "ok" if not errors else ("partial" if actions else "error")
        if status == "ok":
            _upsert_alert(
                source="scheduler",
                code="RESTART_EXECUTED",
                severity="info",
                title="Reinicio ejecutado",
                details=f"target={t} reason={reason}",
                data={"target": t, "reason": reason, "queue": queue_stats},
            )
            _resolve_alert_code("RESTART_POSTPONED_QUEUE")
        else:
            _upsert_alert(
                source="scheduler",
                code="RESTART_ERROR",
                severity="critical",
                title="Fallo en reinicio",
                details="; ".join(errors),
                data={"target": t, "reason": reason, "actions": actions},
            )
        return {
            "status": status,
            "target": t,
            "reason": reason,
            "actions": actions,
            "errors": errors,
            "queue": queue_stats,
        }


def _schedule_next_restart() -> None:
    snap = _restart_settings_snapshot()
    if not snap["enabled"]:
        _setting_set("auto_restart_next_at", "")
        return
    next_ts = time.time() + (int(snap["interval_minutes"]) * 60)
    _setting_set("auto_restart_next_at", _uy_from_epoch_iso(next_ts))


def _restart_worker_loop():
    while True:
        try:
            snap = _restart_settings_snapshot()
            if not snap["enabled"]:
                time.sleep(1.0)
                continue
            remaining = snap["seconds_remaining"]
            if remaining is None or remaining > 0:
                time.sleep(1.0)
                continue
            result = _perform_controlled_restart(snap["target"], reason="scheduled")
            if result.get("status") != "postponed":
                _schedule_next_restart()
            time.sleep(1.0)
        except Exception:
            time.sleep(2.0)


def _upsert_alert(*, source: str, code: str, severity: str, title: str, details: str, data: dict | None = None) -> int:
    sev = str(severity or "warning").strip().lower()
    if sev not in ALERT_SEVERITY_RANK:
        sev = "warning"
    payload = json.dumps(data or {}, ensure_ascii=True, default=str)
    now = _utc_now_iso()

    def _insert_alert_event(
        conn: sqlite3.Connection,
        *,
        alert_id: int,
        event_type: str,
        status: str,
        is_active: bool,
        occurrences: int,
        first_seen: str,
        last_seen: str,
        resolved_at: str | None,
    ) -> None:
        conn.execute(
            """
            INSERT INTO alerts_events (
                ts, alert_id, source, code, severity, title, details, data_json,
                event_type, status, is_active, occurrences, first_seen, last_seen, resolved_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                now,
                int(alert_id),
                str(source),
                str(code),
                sev,
                str(title),
                str(details),
                payload,
                str(event_type),
                str(status),
                1 if bool(is_active) else 0,
                int(max(1, occurrences)),
                str(first_seen or ""),
                str(last_seen or ""),
                str(resolved_at or "") if resolved_at else None,
            ),
        )

    with alerts_lock:
        with _db_conn() as conn:
            row = conn.execute(
                """
                SELECT id, occurrences, severity, title, details, data_json, first_seen
                FROM alerts_log
                WHERE code = ? AND is_active = 1
                ORDER BY id DESC
                LIMIT 1
                """,
                (str(code),),
            ).fetchone()
            if row:
                next_occ = int(row["occurrences"] or 0) + 1
                conn.execute(
                    """
                    UPDATE alerts_log
                    SET ts = ?, source = ?, severity = ?, title = ?, details = ?, data_json = ?,
                        last_seen = ?, occurrences = ?, resolved_at = NULL
                    WHERE id = ?
                    """,
                    (
                        now,
                        str(source),
                        sev,
                        str(title),
                        str(details),
                        payload,
                        now,
                        next_occ,
                        int(row["id"]),
                    ),
                )
                changed = (
                    str(row["severity"] or "") != sev
                    or str(row["title"] or "") != str(title)
                    or str(row["details"] or "") != str(details)
                    or str(row["data_json"] or "") != payload
                )
                if changed:
                    _insert_alert_event(
                        conn,
                        alert_id=int(row["id"]),
                        event_type="update",
                        status="ACTIVA",
                        is_active=True,
                        occurrences=next_occ,
                        first_seen=str(row["first_seen"] or now),
                        last_seen=now,
                        resolved_at=None,
                    )
                conn.commit()
                return int(row["id"])
            conn.execute(
                """
                INSERT INTO alerts_log (
                    ts, source, code, severity, title, details, data_json, is_active,
                    first_seen, last_seen, resolved_at, occurrences, notified_discord
                ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, ?, NULL, 1, 0)
                """,
                (now, str(source), str(code), sev, str(title), str(details), payload, now, now),
            )
            alert_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            _insert_alert_event(
                conn,
                alert_id=alert_id,
                event_type="active",
                status="ACTIVA",
                is_active=True,
                occurrences=1,
                first_seen=now,
                last_seen=now,
                resolved_at=None,
            )
            conn.commit()
            return alert_id


def _resolve_alert_code(code: str) -> None:
    now = _utc_now_iso()
    with alerts_lock:
        with _db_conn() as conn:
            active_rows = conn.execute(
                """
                SELECT id, source, code, severity, title, details, data_json, occurrences, first_seen, last_seen
                FROM alerts_log
                WHERE code = ? AND is_active = 1
                """,
                (str(code),),
            ).fetchall()
            conn.execute(
                """
                UPDATE alerts_log
                SET is_active = 0, resolved_at = ?, last_seen = ?
                WHERE code = ? AND is_active = 1
                """,
                (now, now, str(code)),
            )
            for r in active_rows:
                conn.execute(
                    """
                    INSERT INTO alerts_events (
                        ts, alert_id, source, code, severity, title, details, data_json,
                        event_type, status, is_active, occurrences, first_seen, last_seen, resolved_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'resolved', 'RESUELTA', 0, ?, ?, ?, ?)
                    """,
                    (
                        now,
                        int(r["id"]),
                        str(r["source"] or ""),
                        str(r["code"] or ""),
                        str(r["severity"] or "warning"),
                        str(r["title"] or ""),
                        str(r["details"] or ""),
                        str(r["data_json"] or ""),
                        int(r["occurrences"] or 1),
                        str(r["first_seen"] or ""),
                        str(r["last_seen"] or now),
                        now,
                    ),
                )
            conn.commit()


def _alert_check_or_resolve(active: bool, *, source: str, code: str, severity: str, title: str, details: str, data: dict | None = None):
    if active:
        _upsert_alert(source=source, code=code, severity=severity, title=title, details=details, data=data)
    else:
        _resolve_alert_code(code)


def _severity_allowed(sev: str, min_sev: str) -> bool:
    return ALERT_SEVERITY_RANK.get(str(sev), 0) >= ALERT_SEVERITY_RANK.get(str(min_sev), 0)


def _send_discord_message(webhook_url: str, content: str) -> bool:
    try:
        data = json.dumps({"content": str(content)}, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            str(webhook_url),
            data=data,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=8) as _:
            return True
    except Exception:
        return False


def _dispatch_pending_discord_alerts():
    cfg = _alerts_settings_snapshot()
    if not cfg["discord_enabled"] or not cfg["discord_webhook_url"]:
        return
    with alerts_lock:
        with _db_conn() as conn:
            rows = conn.execute(
                """
                SELECT id, ts, source, code, severity, title, details
                FROM alerts_log
                WHERE is_active = 1 AND notified_discord = 0
                ORDER BY id ASC
                LIMIT 10
                """
            ).fetchall()
            for r in rows:
                sev = str(r["severity"] or "warning")
                if not _severity_allowed(sev, cfg["discord_min_severity"]):
                    conn.execute("UPDATE alerts_log SET notified_discord = 1 WHERE id = ?", (int(r["id"]),))
                    continue
                content = (
                    f"[{sev.upper()}] {r['title']}\n"
                    f"code={r['code']} source={r['source']} ts={r['ts']}\n"
                    f"{r['details']}"
                )
                ok = _send_discord_message(cfg["discord_webhook_url"], content)
                if ok:
                    conn.execute("UPDATE alerts_log SET notified_discord = 1 WHERE id = ?", (int(r["id"]),))
            conn.commit()


def _evaluate_alerts_once():
    cfg = _alerts_settings_snapshot()
    if not cfg["alerts_enabled"]:
        return
    queue_stats = _queue_pending_stats()
    _alert_check_or_resolve(
        int(queue_stats["count"]) >= int(cfg["alerts_queue_pending_threshold"]),
        source="queue",
        code="QUEUE_PENDING_BACKLOG",
        severity="warning",
        title="Backlog en queue/pending",
        details=f"queue_pending={queue_stats['count']} threshold={cfg['alerts_queue_pending_threshold']}",
        data=queue_stats,
    )
    _alert_check_or_resolve(
        int(queue_stats["oldest_age_sec"]) >= int(cfg["alerts_queue_oldest_sec"]),
        source="queue",
        code="QUEUE_PENDING_OLD",
        severity="warning",
        title="Eventos pendientes envejecidos",
        details=f"oldest_age_sec={queue_stats['oldest_age_sec']} threshold={cfg['alerts_queue_oldest_sec']}",
        data=queue_stats,
    )
    _alert_check_or_resolve(
        not lector_manager.running(),
        source="runtime",
        code="LECTOR_OFFLINE",
        severity="warning",
        title="Lector offline",
        details=f"Lector no está ejecutándose (last_exit={lector_manager.last_exit})",
        data={"last_exit": lector_manager.last_exit},
    )
    _alert_check_or_resolve(
        not operador_manager.running(),
        source="runtime",
        code="OPERADOR_OFFLINE",
        severity="critical",
        title="Operador offline",
        details=f"Operador no está ejecutándose (last_exit={operador_manager.last_exit})",
        data={"last_exit": operador_manager.last_exit},
    )
    now = datetime.now(URUGUAY_TZ)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat(timespec="seconds")
    err_since = _uy_from_epoch_iso(time.time() - (int(cfg["alerts_error_window_min"]) * 60))
    stale_cutoff = _uy_from_epoch_iso(time.time() - int(cfg["alerts_stale_sync_sec"]))
    pending_threshold = int(cfg["alerts_pending_order_sec"])
    with _db_conn() as conn:
        recent_errors = int(
            conn.execute(
                """
                SELECT COUNT(1)
                FROM strategy_event_log
                WHERE ts >= ?
                  AND ((error_type IS NOT NULL AND error_type <> '') OR UPPER(COALESCE(status,'')) = 'ERROR')
                """,
                (err_since,),
            ).fetchone()[0]
        )
        no_tickets = int(
            conn.execute(
                """
                SELECT COUNT(1)
                FROM strategy_event_log
                WHERE ts >= ? AND error_type = 'no_tickets_indexed'
                """,
                (err_since,),
            ).fetchone()[0]
        )
        old_pending = int(
            conn.execute(
                """
                SELECT COUNT(1)
                FROM operation_records
                WHERE status = 'PENDING'
                  AND opened_at <> ''
                  AND CAST((julianday('now') - julianday(opened_at)) * 86400 AS INTEGER) >= ?
                """,
                (pending_threshold,),
            ).fetchone()[0]
        )
        stale_sync = int(
            conn.execute(
                """
                SELECT COUNT(1)
                FROM operation_records
                WHERE status IN ('OPEN','PENDING')
                  AND (last_sync_at IS NULL OR last_sync_at = '' OR last_sync_at < ?)
                """,
                (stale_cutoff,),
            ).fetchone()[0]
        )
        day_pnl = float(
            conn.execute(
                """
                SELECT COALESCE(SUM(COALESCE(pnl_usd,0)), 0)
                FROM operation_records
                WHERE status = 'CLOSED' AND closed_at >= ?
                """,
                (today_start,),
            ).fetchone()[0]
        )
    _alert_check_or_resolve(
        recent_errors >= int(cfg["alerts_error_count_threshold"]),
        source="metrics",
        code="ERROR_RATE_HIGH",
        severity="warning",
        title="Tasa de errores alta",
        details=f"errors={recent_errors} en {cfg['alerts_error_window_min']} min (threshold={cfg['alerts_error_count_threshold']})",
        data={"errors": recent_errors, "window_min": cfg["alerts_error_window_min"]},
    )
    _alert_check_or_resolve(
        no_tickets >= int(cfg["alerts_no_tickets_threshold"]),
        source="metrics",
        code="NO_TICKETS_INDEXED_REPEATED",
        severity="warning",
        title="Cierres/modificaciones sin ticket indexado",
        details=f"no_tickets_indexed={no_tickets} en {cfg['alerts_error_window_min']} min",
        data={"count": no_tickets},
    )
    _alert_check_or_resolve(
        old_pending > 0,
        source="metrics",
        code="PENDING_TOO_OLD",
        severity="warning",
        title="Órdenes pendientes demasiado tiempo",
        details=f"pending_old={old_pending} threshold_sec={pending_threshold}",
        data={"pending_old": old_pending},
    )
    _alert_check_or_resolve(
        operador_manager.running() and stale_sync > 0,
        source="metrics",
        code="OPEN_SYNC_STALE",
        severity="warning",
        title="Operaciones sin refresco reciente",
        details=f"open_or_pending_stale={stale_sync} stale_sec={cfg['alerts_stale_sync_sec']}",
        data={"stale_count": stale_sync},
    )
    drawdown_limit = float(cfg["alerts_drawdown_daily_usd"])
    _alert_check_or_resolve(
        day_pnl <= drawdown_limit,
        source="risk",
        code="DAILY_DRAWDOWN_LIMIT",
        severity="critical",
        title="Límite de drawdown diario alcanzado",
        details=f"day_pnl_usd={day_pnl:.2f} limit={drawdown_limit:.2f}",
        data={"day_pnl_usd": day_pnl, "limit": drawdown_limit},
    )


def _alerts_worker_loop():
    while True:
        try:
            cfg = _alerts_settings_snapshot()
            if cfg["alerts_enabled"]:
                _evaluate_alerts_once()
                _dispatch_pending_discord_alerts()
            time.sleep(max(5, int(cfg["alerts_check_interval_sec"])))
        except Exception:
            time.sleep(5.0)


def _alerts_active_items() -> list[dict]:
    _ensure_reports_tables()
    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT id, ts, source, code, severity, title, details, data_json, first_seen, last_seen, occurrences
            FROM alerts_log
            WHERE is_active = 1
            ORDER BY
                CASE severity WHEN 'critical' THEN 0 WHEN 'warning' THEN 1 ELSE 2 END,
                last_seen DESC, id DESC
            """
        ).fetchall()
    return [
        {
            "id": int(r["id"]),
            "ts": str(r["ts"] or ""),
            "source": str(r["source"] or ""),
            "code": str(r["code"] or ""),
            "severity": str(r["severity"] or ""),
            "title": str(r["title"] or ""),
            "details": str(r["details"] or ""),
            "data_json": str(r["data_json"] or ""),
            "first_seen": str(r["first_seen"] or ""),
            "last_seen": str(r["last_seen"] or ""),
            "occurrences": int(r["occurrences"] or 0),
        }
        for r in rows
    ]


def _alerts_history(page: int, page_size: int, from_ts: str | None = None, to_ts: str | None = None) -> dict:
    p = max(1, int(page))
    ps = max(1, min(int(page_size), 20))
    from_dt = _parse_iso_dt_or_none(from_ts)
    to_dt = _parse_iso_dt_or_none(to_ts)
    where = []
    args: list = []
    if from_dt is not None:
        where.append("ts >= ?")
        args.append(from_dt.isoformat(timespec="seconds"))
    if to_dt is not None:
        where.append("ts <= ?")
        args.append(to_dt.isoformat(timespec="seconds"))
    where_sql = (" WHERE " + " AND ".join(where)) if where else ""
    _ensure_reports_tables()
    with _db_conn() as conn:
        total = int(conn.execute(f"SELECT COUNT(1) FROM alerts_events{where_sql}", tuple(args)).fetchone()[0])
        if total <= 0:
            legacy_rows = conn.execute(
                """
                SELECT id, ts, source, code, severity, title, details, data_json, is_active, first_seen, last_seen, resolved_at, occurrences
                FROM alerts_log
                ORDER BY ts DESC, id DESC
                LIMIT 20
                """
            ).fetchall()
            legacy_items = []
            for r in legacy_rows:
                ts_raw = str(r["ts"] or "")
                ts_dt = _parse_iso_dt_or_none(ts_raw)
                if from_dt is not None and ts_dt is not None and ts_dt < from_dt:
                    continue
                if to_dt is not None and ts_dt is not None and ts_dt > to_dt:
                    continue
                legacy_items.append(
                    {
                        "id": int(r["id"]),
                        "alert_id": int(r["id"]),
                        "ts": ts_raw,
                        "source": str(r["source"] or ""),
                        "code": str(r["code"] or ""),
                        "severity": str(r["severity"] or ""),
                        "title": str(r["title"] or ""),
                        "details": str(r["details"] or ""),
                        "data_json": str(r["data_json"] or ""),
                        "event_type": "legacy",
                        "status": "ACTIVA" if bool(r["is_active"]) else "RESUELTA",
                        "is_active": bool(r["is_active"]),
                        "first_seen": str(r["first_seen"] or ""),
                        "last_seen": str(r["last_seen"] or ""),
                        "resolved_at": str(r["resolved_at"] or ""),
                        "occurrences": int(r["occurrences"] or 0),
                    }
                )
            return {
                "items": legacy_items,
                "page": 1,
                "page_size": len(legacy_items),
                "total": len(legacy_items),
                "total_pages": 1,
                "has_prev": False,
                "has_next": False,
            }
        total_pages = max(1, (total + ps - 1) // ps)
        if p > total_pages:
            p = total_pages
        start = (p - 1) * ps
        rows = conn.execute(
            f"""
            SELECT id, alert_id, ts, source, code, severity, title, details, data_json, event_type,
                   status, is_active, first_seen, last_seen, resolved_at, occurrences
            FROM alerts_events
            {where_sql}
            ORDER BY ts DESC, id DESC
            LIMIT ? OFFSET ?
            """,
            tuple(args + [ps, start]),
        ).fetchall()
    items = [
        {
            "id": int(r["id"]),
            "alert_id": int(r["alert_id"] or 0) if r["alert_id"] is not None else 0,
            "ts": str(r["ts"] or ""),
            "source": str(r["source"] or ""),
            "code": str(r["code"] or ""),
            "severity": str(r["severity"] or ""),
            "title": str(r["title"] or ""),
            "details": str(r["details"] or ""),
            "data_json": str(r["data_json"] or ""),
            "event_type": str(r["event_type"] or ""),
            "status": str(r["status"] or ""),
            "is_active": bool(r["is_active"]),
            "first_seen": str(r["first_seen"] or ""),
            "last_seen": str(r["last_seen"] or ""),
            "resolved_at": str(r["resolved_at"] or ""),
            "occurrences": int(r["occurrences"] or 0),
        }
        for r in rows
    ]
    return {
        "items": items,
        "page": p,
        "page_size": ps,
        "total": total,
        "total_pages": total_pages,
        "has_prev": p > 1,
        "has_next": p < total_pages,
    }


def _retention_settings_snapshot() -> dict:
    raw = _settings_values()
    return {
        "enabled": _setting_get_bool(raw, "retention_enabled", True),
        "interval_min": max(5, _setting_get_int(raw, "retention_run_interval_min", 60)),
        "archive_enabled": _setting_get_bool(raw, "retention_archive_enabled", True),
        "strategy_days": max(7, _setting_get_int(raw, "retention_strategy_days", 180)),
        "operation_events_days": max(7, _setting_get_int(raw, "retention_operation_events_days", 180)),
        "telegram_messages_days": max(30, _setting_get_int(raw, "retention_telegram_messages_days", 365)),
        "alerts_events_days": max(30, _setting_get_int(raw, "retention_alerts_events_days", 365)),
        "processed_events_days": max(7, _setting_get_int(raw, "retention_processed_events_days", 90)),
        "last_run_at": str(raw.get("retention_last_run_at", "") or "").strip(),
    }


def _ensure_retention_archive_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS retention_archive (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_name TEXT NOT NULL,
            src_key TEXT NOT NULL,
            archived_at TEXT NOT NULL,
            row_json TEXT NOT NULL
        )
        """
    )
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_retention_archive_src ON retention_archive(table_name, src_key)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_retention_archive_ts ON retention_archive(archived_at)")


def _retention_cleanup_table(
    conn: sqlite3.Connection,
    *,
    table_name: str,
    ts_col: str,
    pk_col: str,
    cutoff_iso: str,
    archive_enabled: bool,
    batch_size: int = 1000,
) -> int:
    rows = conn.execute(
        f"""
        SELECT *
        FROM {table_name}
        WHERE COALESCE({ts_col}, '') <> ''
          AND {ts_col} < ?
        ORDER BY {ts_col} ASC
        LIMIT ?
        """,
        (cutoff_iso, int(batch_size)),
    ).fetchall()
    if not rows:
        return 0
    archived_at = _utc_now_iso()
    keys: list[str] = []
    if archive_enabled:
        _ensure_retention_archive_table(conn)
    for r in rows:
        key = str(r[pk_col] if pk_col in r.keys() else "")
        keys.append(key)
        if archive_enabled:
            payload = json.dumps({k: r[k] for k in r.keys()}, ensure_ascii=True, default=str)
            conn.execute(
                """
                INSERT OR IGNORE INTO retention_archive (table_name, src_key, archived_at, row_json)
                VALUES (?, ?, ?, ?)
                """,
                (str(table_name), key, archived_at, payload),
            )
    if not keys:
        return 0
    ph = ",".join(["?"] * len(keys))
    conn.execute(
        f"DELETE FROM {table_name} WHERE {pk_col} IN ({ph})",
        tuple(keys),
    )
    return len(keys)


def _run_retention_once() -> dict:
    cfg = _retention_settings_snapshot()
    if not cfg["enabled"]:
        return {"status": "disabled", "deleted": 0}
    now = datetime.now(URUGUAY_TZ)
    targets = [
        ("strategy_event_log", "ts", "id", int(cfg["strategy_days"])),
        ("operation_events", "ts", "id", int(cfg["operation_events_days"])),
        ("telegram_messages", "ts", "id", int(cfg["telegram_messages_days"])),
        ("alerts_events", "ts", "id", int(cfg["alerts_events_days"])),
        ("processed_events", "updated_at", "event_uid", int(cfg["processed_events_days"])),
        ("event_retry_state", "updated_at", "event_uid", int(cfg["processed_events_days"])),
        ("queue_event_failures", "last_seen", "event_path", int(cfg["processed_events_days"])),
    ]
    summary: dict[str, int] = {}
    deleted_total = 0
    with _db_conn() as conn:
        for table_name, ts_col, pk_col, keep_days in targets:
            cutoff = (now - timedelta(days=max(1, int(keep_days)))).isoformat(timespec="seconds")
            table_deleted = 0
            while True:
                try:
                    deleted = _retention_cleanup_table(
                        conn,
                        table_name=table_name,
                        ts_col=ts_col,
                        pk_col=pk_col,
                        cutoff_iso=cutoff,
                        archive_enabled=bool(cfg["archive_enabled"]),
                    )
                except Exception:
                    deleted = 0
                if deleted <= 0:
                    break
                table_deleted += int(deleted)
                if deleted < 1000:
                    break
            if table_deleted > 0:
                summary[table_name] = table_deleted
                deleted_total += table_deleted
        conn.commit()
    _setting_set("retention_last_run_at", _utc_now_iso())
    return {"status": "ok", "deleted": int(deleted_total), "tables": summary}


def _retention_worker_loop():
    while True:
        try:
            cfg = _retention_settings_snapshot()
            if cfg["enabled"]:
                due = True
                last_dt = _parse_iso_utc_or_none(cfg["last_run_at"])
                if last_dt is not None:
                    now_dt = datetime.now(URUGUAY_TZ).replace(tzinfo=None)
                    elapsed = (now_dt - last_dt).total_seconds()
                    due = elapsed >= (int(cfg["interval_min"]) * 60)
                if due:
                    _run_retention_once()
            time.sleep(30)
        except Exception:
            time.sleep(30)


def _start_background_workers_once() -> None:
    global workers_started
    if workers_started:
        return
    workers_started = True
    threading.Thread(target=_restart_worker_loop, daemon=True, name="restart-worker").start()
    threading.Thread(target=_alerts_worker_loop, daemon=True, name="alerts-worker").start()
    threading.Thread(target=_retention_worker_loop, daemon=True, name="retention-worker").start()


@app.middleware("http")
async def require_basic_auth(request: Request, call_next):
    # Salud local mínima sin auth estricta para facilitar diagnosis local.
    if request.url.path == "/healthz":
        return JSONResponse({"status": "ok"})
    if not _is_auth_valid(request):
        return Response(
            status_code=401,
            content="Unauthorized",
            headers={"WWW-Authenticate": 'Basic realm="TradingBot"'},
            media_type="text/plain",
        )
    return await call_next(request)


_ensure_channels_table()
_ensure_execution_profiles_table()
_ensure_operator_presets_table()
_ensure_assignments_table()
_ensure_channel_preset_events_table()
_ensure_reports_tables()
_ensure_app_settings_table()
_ensure_runtime_env_table()
_seed_recommended_presets()
_migrate_profiles_and_assignments_once()
_seed_channel_preset_events_for_existing_assignments()
_restore_process_manager_env_cache()
_start_background_workers_once()


def _sse_stream(hub: LogHub):
    q, snapshot = hub.subscribe()
    try:
        for line in snapshot:
            yield f"data: {line}\n\n"
        while True:
            try:
                line = q.get(timeout=1)
                yield f"data: {line}\n\n"
            except queue.Empty:
                yield ": ping\n\n"
    finally:
        hub.unsubscribe(q)


@app.get("/")
def index():
    html = (STATIC_DIR / "index.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/presets")
def presets_page():
    html = (STATIC_DIR / "configs.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/configs")
def configs_page_legacy():
    return RedirectResponse(url="/presets", status_code=307)


@app.get("/reportes")
def reportes_page():
    html = (STATIC_DIR / "reportes.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/alertas")
def alertas_page():
    html = (STATIC_DIR / "alertas.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/canal-presets")
def canal_presets_page():
    html = (STATIC_DIR / "canal_preset_detalle.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/operaciones")
def operaciones_page():
    html = (STATIC_DIR / "operaciones.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/operaciones/historial")
def operaciones_historial_page():
    html = (STATIC_DIR / "operaciones_historial.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/operaciones/{operation_id}")
def operation_detail_page(operation_id: int):
    _ = operation_id
    html = (STATIC_DIR / "operacion_detalle.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/mensajes")
def mensajes_page():
    html = (STATIC_DIR / "mensajes.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/tutorial")
def tutorial_page():
    html = (STATIC_DIR / "tutorial.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/api/status")
def status():
    presets = _list_operator_presets()
    assignments = _list_assignments()
    profiles = _list_execution_profiles()
    restart_cfg = _restart_settings_snapshot()
    queue_stats = _queue_pending_stats()
    with _db_conn() as conn:
        open_ops = int(
            conn.execute(
                "SELECT COUNT(1) FROM operation_records WHERE status IN ('OPEN','PENDING')"
            ).fetchone()[0]
        )
        msgs = int(conn.execute("SELECT COUNT(1) FROM telegram_messages").fetchone()[0])
        alerts_active = int(conn.execute("SELECT COUNT(1) FROM alerts_log WHERE is_active = 1").fetchone()[0])
    return JSONResponse(
        {
            "db_path": str(DB_PATH),
            "mt5_terminal_default": MT5_TERMINAL_DEFAULT,
            "operador_defaults": _get_operator_defaults(),
            "auto_restart": restart_cfg,
            "queue_pending": queue_stats,
            "counts": {
                "channels": len(_list_channels()),
                "presets": len(presets),
                "assignments": len(assignments),
                "profiles": len(profiles),
                "operations_open": open_ops,
                "messages": msgs,
                "alerts_active": alerts_active,
            },
            "lector": {
                "running": lector_manager.running(),
                "pid": lector_manager.proc.pid if lector_manager.proc else None,
                "last_exit": lector_manager.last_exit,
            },
            "operador": {
                "running": operador_manager.running(),
                "pid": operador_manager.proc.pid if operador_manager.proc else None,
                "last_exit": operador_manager.last_exit,
            },
        }
    )


@app.post("/api/web-auth/verify-password")
def verify_web_password(payload: WebAuthPasswordPayload):
    pwd = str(payload.password or "").strip()
    if not pwd:
        raise HTTPException(status_code=400, detail="password requerido")
    if not _is_web_password_valid(pwd):
        raise HTTPException(status_code=401, detail="Contraseña incorrecta")
    return JSONResponse({"valid": True})


@app.get("/api/channels")
def get_channels():
    return JSONResponse({"channels": _list_channels()})


@app.post("/api/channels")
def create_channel(payload: ChannelCreateRequest):
    name, chat_id, external_id, is_active = _normalize_channel_payload(
        payload.name, payload.chat_id, payload.external_id or "", payload.is_active
    )
    now = _utc_now_iso()
    try:
        with _db_conn() as conn:
            conn.execute(
                """
                INSERT INTO telegram_channels (name, chat_id, external_id, is_active, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (name, chat_id, external_id, is_active, now, now),
            )
            new_channel_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            _sync_channel_preset_assignments_conn(
                conn,
                reason="channel_created",
                metadata={"channel_id": new_channel_id},
                now_ts=now,
            )
            conn.commit()
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="Channel name or chat_id already exists") from exc
    return JSONResponse({"status": "created", "channels": _list_channels()})


@app.put("/api/channels/{channel_id}")
def update_channel(channel_id: int, payload: ChannelUpdateRequest):
    name, chat_id, external_id, is_active = _normalize_channel_payload(
        payload.name, payload.chat_id, payload.external_id or "", payload.is_active
    )
    now = _utc_now_iso()
    try:
        with _db_conn() as conn:
            row = conn.execute("SELECT id FROM telegram_channels WHERE id = ?", (channel_id,)).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Channel not found")
            conn.execute(
                """
                UPDATE telegram_channels
                SET name = ?, chat_id = ?, external_id = ?, is_active = ?, updated_at = ?
                WHERE id = ?
                """,
                (name, chat_id, external_id, is_active, now, channel_id),
            )
            _sync_channel_preset_assignments_conn(
                conn,
                reason="channel_updated",
                metadata={"channel_id": int(channel_id)},
                now_ts=now,
            )
            conn.commit()
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="Channel name or chat_id already exists") from exc
    return JSONResponse({"status": "updated", "channels": _list_channels()})


@app.delete("/api/channels/{channel_id}")
def delete_channel(channel_id: int):
    with _db_conn() as conn:
        row = conn.execute("SELECT id FROM telegram_channels WHERE id = ?", (channel_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Channel not found")
        assignment_rows = conn.execute(
            """
            SELECT
                a.id AS assignment_id, a.channel_id, a.config_id, a.mode, a.is_active,
                c.name AS channel_name, c.chat_id AS channel_chat_id,
                p.name AS config_name
            FROM channel_config_assignments a
            LEFT JOIN telegram_channels c ON c.id = a.channel_id
            LEFT JOIN operator_presets p ON p.id = a.config_id
            WHERE a.channel_id = ?
            """,
            (int(channel_id),),
        ).fetchall()
        now = _utc_now_iso()
        for a in assignment_rows:
            snap = {
                "channel_id": a["channel_id"],
                "channel_name": a["channel_name"],
                "channel_chat_id": a["channel_chat_id"],
                "config_id": a["config_id"],
                "config_name": a["config_name"],
                "mode": a["mode"],
                "is_active": 0,
            }
            _append_channel_preset_event_conn(
                conn,
                assignment_id=int(a["assignment_id"]),
                event_type="deleted",
                snapshot=snap,
                details="assignment_deleted_by_channel_delete",
                metadata={"source": "api", "reason": "channel_deleted"},
                ts=now,
            )
        conn.execute("DELETE FROM channel_config_assignments WHERE channel_id = ?", (channel_id,))
        conn.execute("DELETE FROM telegram_channels WHERE id = ?", (channel_id,))
        _sync_channel_preset_assignments_conn(
            conn,
            reason="channel_deleted",
            metadata={"channel_id": int(channel_id)},
            now_ts=now,
        )
        conn.commit()
    return JSONResponse({"status": "deleted", "channels": _list_channels()})


@app.get("/api/operator-presets")
def list_operator_presets():
    return JSONResponse({"presets": _list_operator_presets()})


@app.get("/api/execution-profiles")
def list_execution_profiles():
    return JSONResponse({"profiles": _list_execution_profiles()})


@app.post("/api/execution-profiles")
def create_execution_profile(payload: ExecutionProfilePayload):
    data = _normalize_execution_profile_payload(payload)
    now = _utc_now_iso()
    try:
        with _db_conn() as conn:
            conn.execute(
                """
                INSERT INTO execution_profiles (code, name, description, is_system, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (data["code"], data["name"], data["description"], data["is_system"], now, now),
            )
            conn.commit()
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="Profile code already exists") from exc
    return JSONResponse({"status": "created", "profiles": _list_execution_profiles()})


@app.put("/api/execution-profiles/{profile_id}")
def update_execution_profile(profile_id: int, payload: ExecutionProfilePayload):
    data = _normalize_execution_profile_payload(payload)
    now = _utc_now_iso()
    try:
        with _db_conn() as conn:
            row = conn.execute("SELECT id, code, is_system FROM execution_profiles WHERE id = ?", (int(profile_id),)).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Profile not found")
            if int(row["is_system"]) == 1 and data["code"] != str(row["code"]).upper():
                raise HTTPException(status_code=400, detail="System profile code cannot be changed")
            conn.execute(
                """
                UPDATE execution_profiles
                SET code = ?, name = ?, description = ?, is_system = ?, updated_at = ?
                WHERE id = ?
                """,
                (data["code"], data["name"], data["description"], data["is_system"], now, int(profile_id)),
            )
            conn.commit()
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="Profile code already exists") from exc
    return JSONResponse({"status": "updated", "profiles": _list_execution_profiles()})


@app.delete("/api/execution-profiles/{profile_id}")
def delete_execution_profile(profile_id: int):
    with _db_conn() as conn:
        row = conn.execute("SELECT id, is_system FROM execution_profiles WHERE id = ?", (int(profile_id),)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Profile not found")
        if int(row["is_system"]) == 1:
            raise HTTPException(status_code=400, detail="System profile cannot be deleted")
        in_use = conn.execute(
            "SELECT COUNT(1) FROM operator_presets WHERE execution_profile_id = ?",
            (int(profile_id),),
        ).fetchone()[0]
        if int(in_use) > 0:
            raise HTTPException(status_code=409, detail="Profile is in use by presets")
        conn.execute("DELETE FROM execution_profiles WHERE id = ?", (int(profile_id),))
        conn.commit()
    return JSONResponse({"status": "deleted", "profiles": _list_execution_profiles()})


@app.get("/api/configs")
def list_configs_alias():
    return JSONResponse({"configs": _list_operator_presets()})


@app.post("/api/operator-presets")
def create_operator_preset(payload: OperatorPresetPayload):
    data = _normalize_operator_preset_payload(payload)
    now = _utc_now_iso()
    try:
        with _db_conn() as conn:
            prof = conn.execute(
                "SELECT id FROM execution_profiles WHERE id = ?",
                (int(data["execution_profile_id"]),),
            ).fetchone()
            if not prof:
                raise HTTPException(status_code=404, detail="Execution profile not found")
            _validate_single_real_preset_conn(
                conn,
                execution_profile_id=int(data["execution_profile_id"]),
                is_real=bool(data["is_default"]),
            )
            conn.execute(
                """
                INSERT INTO operator_presets (
                    name, mt5_terminal_path, mt5_login, mt5_server,
                    execution_profile_id, total_volume, near_entry_pips_min, near_entry_spread_mult,
                    verify_order_after_send, auto_close_on_mismatch, is_default, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    data["name"], data["mt5_terminal_path"], data["mt5_login"], data["mt5_server"],
                    data["execution_profile_id"], data["total_volume"], data["near_entry_pips_min"], data["near_entry_spread_mult"],
                    data["verify_order_after_send"], data["auto_close_on_mismatch"], data["is_default"], now, now,
                ),
            )
            preset_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            if data["is_default"]:
                _set_default_preset(conn, preset_id)
            _sync_channel_preset_assignments_conn(
                conn,
                reason="preset_created",
                metadata={"preset_id": int(preset_id)},
                now_ts=now,
            )
            conn.commit()
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="Preset name already exists") from exc
    return JSONResponse({"status": "created", "presets": _list_operator_presets()})


@app.put("/api/operator-presets/{preset_id}")
def update_operator_preset(preset_id: int, payload: OperatorPresetPayload):
    data = _normalize_operator_preset_payload(payload)
    now = _utc_now_iso()
    try:
        with _db_conn() as conn:
            row = conn.execute("SELECT id FROM operator_presets WHERE id = ?", (preset_id,)).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Preset not found")
            prof = conn.execute(
                "SELECT id FROM execution_profiles WHERE id = ?",
                (int(data["execution_profile_id"]),),
            ).fetchone()
            if not prof:
                raise HTTPException(status_code=404, detail="Execution profile not found")
            _validate_single_real_preset_conn(
                conn,
                execution_profile_id=int(data["execution_profile_id"]),
                is_real=bool(data["is_default"]),
                current_preset_id=int(preset_id),
            )
            conn.execute(
                """
                UPDATE operator_presets
                SET name = ?, mt5_terminal_path = ?, mt5_login = ?, mt5_server = ?,
                    execution_profile_id = ?, total_volume = ?, near_entry_pips_min = ?, near_entry_spread_mult = ?,
                    verify_order_after_send = ?, auto_close_on_mismatch = ?, is_default = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    data["name"], data["mt5_terminal_path"], data["mt5_login"], data["mt5_server"],
                    data["execution_profile_id"], data["total_volume"], data["near_entry_pips_min"], data["near_entry_spread_mult"],
                    data["verify_order_after_send"], data["auto_close_on_mismatch"], data["is_default"], now, preset_id,
                ),
            )
            if data["is_default"]:
                _set_default_preset(conn, preset_id)
            _sync_channel_preset_assignments_conn(
                conn,
                reason="preset_updated",
                metadata={"preset_id": int(preset_id)},
                now_ts=now,
            )
            conn.commit()
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="Preset name already exists") from exc
    return JSONResponse({"status": "updated", "presets": _list_operator_presets()})


@app.post("/api/operator-presets/{preset_id}/set-default")
def set_operator_preset_default(preset_id: int):
    with _db_conn() as conn:
        row = conn.execute("SELECT id FROM operator_presets WHERE id = ?", (preset_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Preset not found")
        now = _utc_now_iso()
        _set_default_preset(conn, preset_id)
        _sync_channel_preset_assignments_conn(
            conn,
            reason="preset_set_default",
            metadata={"preset_id": int(preset_id)},
            now_ts=now,
        )
        conn.commit()
    return JSONResponse({"status": "default_set", "presets": _list_operator_presets()})


@app.delete("/api/operator-presets/{preset_id}")
def delete_operator_preset(preset_id: int):
    with _db_conn() as conn:
        row = conn.execute("SELECT id, is_default FROM operator_presets WHERE id = ?", (preset_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Preset not found")
        assignment_rows = conn.execute(
            """
            SELECT
                a.id AS assignment_id, a.channel_id, a.config_id, a.mode, a.is_active,
                c.name AS channel_name, c.chat_id AS channel_chat_id,
                p.name AS config_name
            FROM channel_config_assignments a
            LEFT JOIN telegram_channels c ON c.id = a.channel_id
            LEFT JOIN operator_presets p ON p.id = a.config_id
            WHERE a.config_id = ?
            """,
            (int(preset_id),),
        ).fetchall()
        now = _utc_now_iso()
        for a in assignment_rows:
            snap = {
                "channel_id": a["channel_id"],
                "channel_name": a["channel_name"],
                "channel_chat_id": a["channel_chat_id"],
                "config_id": a["config_id"],
                "config_name": a["config_name"],
                "mode": a["mode"],
                "is_active": 0,
            }
            _append_channel_preset_event_conn(
                conn,
                assignment_id=int(a["assignment_id"]),
                event_type="deleted",
                snapshot=snap,
                details="assignment_deleted_by_preset_delete",
                metadata={"source": "api", "reason": "preset_deleted"},
                ts=now,
            )
        conn.execute("DELETE FROM operator_presets WHERE id = ?", (preset_id,))
        if int(row["is_default"]) == 1:
            other = conn.execute("SELECT id FROM operator_presets ORDER BY id ASC LIMIT 1").fetchone()
            if other:
                _set_default_preset(conn, int(other["id"]))
        conn.execute("DELETE FROM channel_config_assignments WHERE config_id = ?", (preset_id,))
        _sync_channel_preset_assignments_conn(
            conn,
            reason="preset_deleted",
            metadata={"preset_id": int(preset_id)},
            now_ts=now,
        )
        conn.commit()
    return JSONResponse({"status": "deleted", "presets": _list_operator_presets()})


@app.get("/api/assignments")
def list_assignments():
    return JSONResponse({"assignments": _list_assignments()})


@app.post("/api/assignments/seed-cross-product")
def seed_assignments_cross_product():
    raise HTTPException(
        status_code=409,
        detail="Asignaciones Canal.Preset se gestionan automáticamente al crear/editar/eliminar canales o presets.",
    )


@app.post("/api/assignments")
def create_assignment(payload: AssignmentCreateRequest):
    raise HTTPException(
        status_code=409,
        detail="Asignaciones Canal.Preset se crean automáticamente desde canales y presets.",
    )


@app.put("/api/assignments/{assignment_id}")
def update_assignment(assignment_id: int, payload: AssignmentUpdateRequest):
    raise HTTPException(
        status_code=409,
        detail="Asignaciones Canal.Preset se actualizan automáticamente al modificar canales o presets.",
    )


@app.delete("/api/assignments/{assignment_id}")
def delete_assignment(assignment_id: int):
    raise HTTPException(
        status_code=409,
        detail="Asignaciones Canal.Preset se eliminan automáticamente desde canales o presets.",
    )


@app.get("/api/channel-presets/registry")
def channel_presets_registry(
    assignment_id: int | None = None,
    from_ts: str = "",
    to_ts: str = "",
):
    items = _channel_preset_registry(
        assignment_id=assignment_id if assignment_id is not None else None,
        from_ts=from_ts or None,
        to_ts=to_ts or None,
    )
    return JSONResponse({"items": items, "count": len(items)})


@app.get("/api/channel-presets/{assignment_id}/detail")
def channel_preset_detail(assignment_id: int, from_ts: str = "", to_ts: str = ""):
    detail = _channel_preset_detail(int(assignment_id), from_ts=from_ts or None, to_ts=to_ts or None)
    if not detail:
        raise HTTPException(status_code=404, detail="Canal.Preset no encontrado")
    return JSONResponse(detail)


@app.post("/api/channel-presets/{assignment_id}/set-active")
def channel_preset_set_active(assignment_id: int, payload: ChannelPresetSetActivePayload):
    aid = int(assignment_id)
    target_active = 1 if bool(payload.is_active) else 0
    now = _utc_now_iso()
    with _db_conn() as conn:
        row = _assignment_row_by_id(conn, aid)
        if not row:
            raise HTTPException(status_code=404, detail="Canal.Preset no está disponible para activar/desactivar.")
        from_active = int(row["is_active"] or 0)
        if target_active == 0:
            open_count = _open_operations_count_for_channel_preset_conn(
                conn,
                int(row["channel_id"]),
                int(row["config_id"]),
            )
            if open_count > 0:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "Canal.Preset tiene operación abierta, por favor ciérrala antes de desactivar este Canal.Preset."
                    ),
                )
        if from_active != target_active:
            ev_type = "activation" if int(target_active) == 1 else "deactivation"
            ev_details = "assignment_activated_by_api" if int(target_active) == 1 else "assignment_deactivated_by_api"
            conn.execute(
                """
                UPDATE channel_config_assignments
                SET is_active = ?, updated_at = ?
                WHERE id = ?
                """,
                (int(target_active), now, aid),
            )
            snap = _assignment_snapshot_by_id(conn, aid)
            _append_channel_preset_event_conn(
                conn,
                assignment_id=aid,
                event_type=ev_type,
                snapshot=snap,
                details=ev_details,
                metadata={
                    "source": "api",
                    "action": "set_active",
                    "from_is_active": int(from_active),
                    "to_is_active": int(target_active),
                },
                ts=now,
            )
            conn.commit()
    return JSONResponse(
        {
            "status": "updated",
            "assignment_id": aid,
            "is_active": bool(target_active),
        }
    )


@app.get("/api/operations/open")
def operations_open():
    return JSONResponse(
        {
            "items": _fetch_open_operations(),
            "server_ts": _utc_now_iso(),
        }
    )


@app.get("/api/operations/closed")
def operations_closed(
    page: int = 1,
    page_size: int = 50,
    from_ts: str = "",
    to_ts: str = "",
    opened_from_ts: str = "",
    opened_to_ts: str = "",
    closed_from_ts: str = "",
    closed_to_ts: str = "",
    symbol: str = "",
    side: str = "",
    operation_id: str = "",
    channel_preset: str = "",
    close_source: str = "",
    close_message_id: str = "",
    error_id: str = "",
):
    op_id = None
    op_id_raw = str(operation_id or "").strip()
    if op_id_raw:
        try:
            op_id = int(op_id_raw)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="operation_id must be numeric") from exc
    # Compatibilidad: from_ts/to_ts antiguos se interpretan como rango de cierre.
    closed_from_q = str(closed_from_ts or "").strip() or str(from_ts or "").strip()
    closed_to_q = str(closed_to_ts or "").strip() or str(to_ts or "").strip()
    return JSONResponse(
        _fetch_closed_operations(
            page=page,
            page_size=page_size,
            opened_from_ts=opened_from_ts or None,
            opened_to_ts=opened_to_ts or None,
            closed_from_ts=closed_from_q or None,
            closed_to_ts=closed_to_q or None,
            symbol=symbol or None,
            side=side or None,
            operation_id=op_id,
            channel_preset=channel_preset or None,
            close_source=close_source or None,
            close_message_id=close_message_id or None,
            error_id=error_id or None,
        )
    )


@app.get("/api/operations/search")
def operations_search(operation_id: str = ""):
    q = str(operation_id or "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="operation_id is required")
    try:
        op_id = int(q)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="operation_id must be numeric") from exc
    detail = _fetch_operation_detail(op_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Operation not found")
    return JSONResponse({"operation_id": op_id, "url": f"/operaciones/{op_id}"})


@app.get("/api/operations/{operation_id}")
def operation_detail(operation_id: int):
    detail = _fetch_operation_detail(operation_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Operation not found")
    return JSONResponse(detail)


@app.post("/api/operations/{operation_id}/close-manual")
def operation_close_manual(operation_id: int, payload: OperationManualClosePayload):
    result = _close_operation_manual(
        int(operation_id),
        reason=payload.reason,
        details=payload.details,
        close_in_mt5=bool(payload.close_in_mt5),
    )
    return JSONResponse({"status": "closed", **result})


@app.post("/api/operations/close-manual")
def operations_close_manual(payload: OperationsManualClosePayload):
    result = _close_operations_manual_bulk(
        mode=str(payload.mode or "all"),
        include_pending=bool(payload.include_pending),
        reason=payload.reason,
        details=payload.details,
        close_in_mt5=bool(payload.close_in_mt5),
    )
    return JSONResponse({"status": "closed_bulk", **result})


@app.get("/api/messages/search")
def messages_search(message_id: str = "", limit: int = 100):
    q = str(message_id or "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="message_id is required")
    items = _search_messages(q, limit=limit)
    return JSONResponse({"items": items, "count": len(items), "query": q})


@app.get("/api/messages/{message_uid}")
def message_detail(message_uid: str):
    items = _search_messages(str(message_uid or "").strip(), limit=1)
    if not items:
        raise HTTPException(status_code=404, detail="Message not found")
    return JSONResponse({"item": items[0]})


@app.get("/api/restart/config")
def restart_config_get():
    return JSONResponse({"config": _restart_settings_snapshot()})


@app.put("/api/restart/config")
def restart_config_put(payload: RestartConfigPayload):
    config = _set_restart_config(
        enabled=bool(payload.enabled),
        interval_minutes=int(payload.interval_minutes),
        target=str(payload.target or "operador"),
    )
    return JSONResponse({"status": "updated", "config": config})


@app.post("/api/restart/now")
def restart_now(payload: RestartNowPayload):
    snap = _restart_settings_snapshot()
    target = str(payload.target or snap["target"] or "operador").strip().lower()
    result = _perform_controlled_restart(target=target, reason="manual_quick")
    if result.get("status") != "postponed" and snap["enabled"]:
        _schedule_next_restart()
    return JSONResponse({"status": "ok", "result": result, "config": _restart_settings_snapshot()})


@app.get("/api/alerts/config")
def alerts_config_get():
    return JSONResponse({"config": _alerts_settings_snapshot()})


@app.put("/api/alerts/config")
def alerts_config_put(payload: AlertsConfigPayload):
    min_sev = str(payload.discord_min_severity or "warning").strip().lower()
    if min_sev not in ALERT_SEVERITY_RANK:
        raise HTTPException(status_code=400, detail="discord_min_severity inválido")
    _setting_set("alerts_enabled", "true" if bool(payload.alerts_enabled) else "false")
    _setting_set("alerts_check_interval_sec", str(max(5, int(payload.alerts_check_interval_sec))))
    _setting_set("alerts_queue_pending_threshold", str(max(1, int(payload.alerts_queue_pending_threshold))))
    _setting_set("alerts_queue_oldest_sec", str(max(30, int(payload.alerts_queue_oldest_sec))))
    _setting_set("alerts_pending_order_sec", str(max(60, int(payload.alerts_pending_order_sec))))
    _setting_set("alerts_error_window_min", str(max(1, int(payload.alerts_error_window_min))))
    _setting_set("alerts_error_count_threshold", str(max(1, int(payload.alerts_error_count_threshold))))
    _setting_set("alerts_no_tickets_threshold", str(max(1, int(payload.alerts_no_tickets_threshold))))
    _setting_set("alerts_drawdown_daily_usd", str(float(payload.alerts_drawdown_daily_usd)))
    _setting_set("alerts_stale_sync_sec", str(max(30, int(payload.alerts_stale_sync_sec))))
    _setting_set("discord_enabled", "true" if bool(payload.discord_enabled) else "false")
    _setting_set("discord_webhook_url", str(payload.discord_webhook_url or "").strip())
    _setting_set("discord_min_severity", min_sev)
    return JSONResponse({"status": "updated", "config": _alerts_settings_snapshot()})


@app.get("/api/alerts/active")
def alerts_active():
    return JSONResponse({"items": _alerts_active_items()})


@app.get("/api/alerts/history")
def alerts_history(page: int = 1, page_size: int = 20, from_ts: str = "", to_ts: str = ""):
    return JSONResponse(_alerts_history(page=page, page_size=page_size, from_ts=from_ts or None, to_ts=to_ts or None))


@app.post("/api/alerts/discord-test")
def alerts_discord_test():
    cfg = _alerts_settings_snapshot()
    if not cfg["discord_enabled"] or not cfg["discord_webhook_url"]:
        raise HTTPException(status_code=400, detail="Discord webhook no configurado o desactivado")
    content = f"[INFO] Test Discord bot\n{_utc_now_iso()} | source=webapp"
    ok = _send_discord_message(cfg["discord_webhook_url"], content)
    if not ok:
        raise HTTPException(status_code=502, detail="No se pudo enviar al webhook de Discord")
    return JSONResponse({"status": "sent"})


@app.post("/api/operator-presets/seed-recommended")
def operator_presets_seed_recommended():
    _seed_recommended_presets()
    return JSONResponse({"status": "seeded", "presets": _list_operator_presets()})


def _active_combo_lookup_maps(active_combos: dict[str, dict]) -> tuple[dict[str, str], dict[str, str]]:
    by_ids: dict[str, str] = {}
    by_names: dict[str, str] = {}
    for combo_key, item in active_combos.items():
        ch_id = int(item.get("channel_id") or 0)
        cfg_id = int(item.get("config_id") or 0)
        if ch_id > 0 and cfg_id > 0:
            by_ids[f"{ch_id}:{cfg_id}"] = combo_key
        channel_name = str(item.get("channel_name") or "unknown")
        preset_name = str(item.get("preset_name") or item.get("config_name") or "default")
        by_names[f"{channel_name}.{preset_name}"] = combo_key
    return by_ids, by_names


@app.get("/api/reportes/summary")
def reportes_summary():
    rows = _fetch_report_rows(limit=10000)
    agg: dict[str, dict] = _active_assignment_combos()
    by_ids, by_names = _active_combo_lookup_maps(agg)
    for r in rows:
        mode = str(r["mode"] or "unknown")
        ch_id = int(r["channel_id"]) if r["channel_id"] is not None else 0
        cfg_id = int(r["config_id"]) if r["config_id"] is not None else 0
        key = ""
        if ch_id > 0 and cfg_id > 0:
            key = by_ids.get(f"{ch_id}:{cfg_id}", "")
        if not key:
            channel_name = str(r["channel_name"] or "unknown")
            preset_name = str(r["config_name"] or "default")
            key = by_names.get(f"{channel_name}.{preset_name}", "")
        if key not in agg:
            continue
        cur = agg[key]
        cur["mode"] = mode or cur["mode"]
        cur["events"] += 1
        ev = str(r["event_type"] or "")
        if ev == "entry":
            cur["entries"] += 1
        elif ev == "modification":
            cur["modifications"] += 1
        elif ev == "close":
            cur["closes"] += 1
        if str(r["error_type"] or "").strip() != "" or str(r["status"] or "").upper() == "ERROR":
            cur["errors"] += 1
        if r["pnl_usd"] is not None:
            cur["pnl_usd"] += float(r["pnl_usd"])
        if r["pnl_pips"] is not None:
            cur["pnl_pips"] += float(r["pnl_pips"])
    items = sorted(agg.values(), key=lambda x: x["pnl_usd"], reverse=True)
    return JSONResponse({"items": items})


@app.get("/api/reportes/timeseries")
def reportes_timeseries(from_ts: str = "", to_ts: str = ""):
    rows = _fetch_report_rows(limit=20000)
    from_dt = _parse_iso_dt_or_none(from_ts or "")
    to_dt = _parse_iso_dt_or_none(to_ts or "")
    if from_dt is not None and to_dt is not None and to_dt < from_dt:
        raise HTTPException(status_code=400, detail="Rango inválido: to_ts debe ser mayor o igual a from_ts")
    active = _active_assignment_combos()
    by_ids, by_names = _active_combo_lookup_maps(active)
    by_combo: dict[str, list] = {k: [] for k in active.keys()}
    points_total = 0
    for r in reversed(rows):
        pnl = r["pnl_usd"]
        if pnl is None:
            continue
        ts_raw = str(r["ts"] or "")
        ts_dt = _parse_iso_dt_or_none(ts_raw)
        if from_dt is not None and ts_dt is not None and ts_dt < from_dt:
            continue
        if to_dt is not None and ts_dt is not None and ts_dt > to_dt:
            continue
        ch_id = int(r["channel_id"]) if r["channel_id"] is not None else 0
        cfg_id = int(r["config_id"]) if r["config_id"] is not None else 0
        combo = ""
        if ch_id > 0 and cfg_id > 0:
            combo = by_ids.get(f"{ch_id}:{cfg_id}", "")
        if not combo:
            combo = by_names.get(
                f"{str(r['channel_name'] or 'unknown')}.{str(r['config_name'] or 'default')}",
                "",
            )
        if combo not in by_combo:
            continue
        by_combo.setdefault(combo, [])
        last = by_combo[combo][-1]["value"] if by_combo[combo] else 0.0
        by_combo[combo].append({"ts": ts_raw, "value": float(last) + float(pnl)})
        points_total += 1
    return JSONResponse(
        {
            "series": by_combo,
            "filter": {
                "from_ts": from_dt.isoformat(timespec="seconds") if from_dt is not None else "",
                "to_ts": to_dt.isoformat(timespec="seconds") if to_dt is not None else "",
            },
            "points_total": int(points_total),
        }
    )


@app.get("/api/reportes/channel-preset-metrics")
def reportes_channel_preset_metrics():
    return JSONResponse({"items": _channel_preset_metrics_for_active()})


def _parse_iso_dt_or_none(value: str | None):
    s = str(value or "").strip()
    if not s:
        return None
    # Soporta inputs datetime-local (YYYY-MM-DDTHH:MM) e ISO completos.
    dt = None
    try:
        dt = datetime.fromisoformat(s)
    except Exception:
        pass
    if dt is None:
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            return None
    # Normaliza a hora Uruguay sin tz para evitar comparaciones aware vs naive.
    if dt.tzinfo is not None:
        return dt.astimezone(URUGUAY_TZ).replace(tzinfo=None)
    return dt


def _collect_error_items(start_ts: str | None = None, end_ts: str | None = None):
    start_dt = _parse_iso_dt_or_none(start_ts)
    end_dt = _parse_iso_dt_or_none(end_ts)

    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT ts, channel_name, config_name, mode, event_type, status, error_type, symbol, details
            FROM strategy_event_log
            WHERE (error_type IS NOT NULL AND error_type <> '')
               OR UPPER(COALESCE(status,'')) = 'ERROR'
            ORDER BY ts DESC, id DESC
            LIMIT 100000
            """
        ).fetchall()

    items = []
    for r in rows:
        ts_raw = str(r["ts"] or "")
        ts_dt = _parse_iso_dt_or_none(ts_raw)
        if start_dt is not None and ts_dt is not None and ts_dt < start_dt:
            continue
        if end_dt is not None and ts_dt is not None and ts_dt > end_dt:
            continue
        items.append(
            {
                "ts": ts_raw,
                "channel_name": str(r["channel_name"] or ""),
                "config_name": str(r["config_name"] or ""),
                "preset_name": str(r["config_name"] or ""),
                "mode": str(r["mode"] or ""),
                "event_type": str(r["event_type"] or ""),
                "status": str(r["status"] or ""),
                "error_type": str(r["error_type"] or ""),
                "symbol": str(r["symbol"] or ""),
                "details": str(r["details"] or ""),
            }
        )

    # Lector non-signals (parser_none/no_events) como errores de ingestión
    try:
        if NON_SIGNALS_CSV_PATH.exists():
            with open(NON_SIGNALS_CSV_PATH, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    reason = str(row.get("reason") or "").strip()
                    if reason == "":
                        continue
                    ts_raw = str(row.get("timestamp") or "")
                    ts_dt = _parse_iso_dt_or_none(ts_raw)
                    if start_dt is not None and ts_dt is not None and ts_dt < start_dt:
                        continue
                    if end_dt is not None and ts_dt is not None and ts_dt > end_dt:
                        continue
                    items.append(
                        {
                            "ts": ts_raw,
                            "channel_name": str(row.get("channel") or ""),
                            "config_name": "parser",
                            "preset_name": "parser",
                            "mode": "lector",
                            "event_type": "ingestion",
                            "status": "ERROR",
                            "error_type": reason,
                            "symbol": "",
                            "details": str(row.get("text") or "")[:300],
                        }
                    )
    except Exception:
        pass

    items.sort(key=lambda x: str(x.get("ts", "")), reverse=True)
    return items


def _build_errors_excel(items: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Errores"

    headers = ["Timestamp", "Canal.Preset", "Modo", "Evento", "Error", "Detalle"]
    ws.append(headers)
    for it in items:
        ws.append(
            [
                str(it.get("ts") or ""),
                f"{it.get('channel_name', '')}.{it.get('config_name', '')}",
                str(it.get("mode") or ""),
                str(it.get("event_type") or ""),
                str(it.get("error_type") or it.get("status") or ""),
                str(it.get("details") or ""),
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{max(1, ws.max_row)}"

    col_widths = [22, 34, 12, 16, 24, 90]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, ws.max_row + 1):
        max_lines = 1
        for col_idx, width in enumerate(col_widths, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            text = str(cell.value or "")
            explicit_lines = text.count("\n") + 1
            wrap_lines = max(1, (len(text) // max(1, int(width - 2))) + 1)
            max_lines = max(max_lines, explicit_lines, wrap_lines)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = min(220, max(18, (max_lines * 14) + 4))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


@app.get("/api/reportes/errors")
def reportes_errors(
    page: int = 1,
    page_size: int = 100,
    from_ts: str = "",
    to_ts: str = "",
):
    p = max(1, int(page))
    ps = max(1, min(int(page_size), 1000))
    items_all = _collect_error_items(start_ts=from_ts or None, end_ts=to_ts or None)
    total = len(items_all)
    total_pages = max(1, (total + ps - 1) // ps)
    if p > total_pages:
        p = total_pages
    start = (p - 1) * ps
    end = start + ps
    return JSONResponse(
        {
            "items": items_all[start:end],
            "page": p,
            "page_size": ps,
            "total": total,
            "total_pages": total_pages,
            "has_prev": p > 1,
            "has_next": p < total_pages,
        }
    )


@app.get("/api/reportes/errors/excel")
def reportes_errors_excel(from_ts: str = "", to_ts: str = ""):
    items = _collect_error_items(start_ts=from_ts or None, end_ts=to_ts or None)
    excel_bytes = _build_errors_excel(items)
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="errores_reportes.xlsx"'},
    )


@app.get("/api/reportes/errors/pdf")
def reportes_errors_pdf_legacy(from_ts: str = "", to_ts: str = ""):
    # Compatibilidad retroactiva para clientes con frontend en cache.
    payload = {}
    if from_ts:
        payload["from_ts"] = from_ts
    if to_ts:
        payload["to_ts"] = to_ts
    query = ("?" + urlencode(payload)) if payload else ""
    return RedirectResponse(url=f"/api/reportes/errors/excel{query}", status_code=307)


@app.post("/api/start/lector")
def start_lector(payload: LectorStartRequest):
    if _active_channels_count() <= 0:
        raise HTTPException(status_code=400, detail="No hay canales activos en SQLite. Carga al menos uno.")

    api_hash = payload.telegram_api_hash.strip()
    openai_key = payload.openai_api_key.strip()
    openai_model = payload.openai_model.strip()
    if int(payload.telegram_api_id) <= 0:
        raise HTTPException(status_code=400, detail="TELEGRAM_API_ID debe ser un entero positivo")
    if not api_hash:
        raise HTTPException(status_code=400, detail="TELEGRAM_API_HASH es obligatorio")
    if not openai_key:
        raise HTTPException(status_code=400, detail="OPENAI_API_KEY es obligatorio")
    if not openai_model:
        raise HTTPException(status_code=400, detail="OPENAI_MODEL es obligatorio")

    env = os.environ.copy()
    env["PYTHONUNBUFFERED"] = "1"
    env["TELEGRAM_API_ID"] = str(payload.telegram_api_id)
    env["TELEGRAM_API_HASH"] = api_hash
    env["OPENAI_API_KEY"] = openai_key
    env["OPENAI_MODEL"] = openai_model
    if payload.openai_base_url:
        env["OPENAI_BASE_URL"] = payload.openai_base_url.strip()
    else:
        env.pop("OPENAI_BASE_URL", None)
    env["TRADING_BOT_DB_PATH"] = str(DB_PATH)

    try:
        lector_manager.start(env)
        _persist_process_runtime_env(PROCESS_NAME_LECTOR, env)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return {"status": "started"}


@app.post("/api/stop/lector")
def stop_lector():
    lector_manager.stop()
    return {"status": "stopped"}


@app.post("/api/start/operador")
def start_operador(payload: OperadorStartRequest):
    op_defaults = _get_operator_defaults()
    terminal_path = str(payload.mt5_terminal_path or "").strip() or str(op_defaults.get("mt5_terminal_path") or MT5_TERMINAL_DEFAULT)
    mt5_login = int(payload.mt5_login if payload.mt5_login is not None else (op_defaults.get("mt5_login") or 0))
    mt5_password = payload.mt5_password.strip()
    mt5_server = str(payload.mt5_server or "").strip() or str(op_defaults.get("mt5_server") or "")
    execution_profile = str(op_defaults.get("execution_profile_code") or DEFAULT_PROFILE_CODE).strip().upper()
    total_volume = float(payload.total_volume if payload.total_volume is not None else op_defaults.get("total_volume", 0.03))
    near_entry_pips_min = float(payload.near_entry_pips_min if payload.near_entry_pips_min is not None else op_defaults.get("near_entry_pips_min", 1.0))
    near_entry_spread_mult = float(payload.near_entry_spread_mult if payload.near_entry_spread_mult is not None else op_defaults.get("near_entry_spread_mult", 2.0))
    verify_after_send = bool(payload.verify_order_after_send) if payload.verify_order_after_send is not None else bool(op_defaults.get("verify_order_after_send", True))
    auto_close_mismatch = bool(payload.auto_close_on_mismatch) if payload.auto_close_on_mismatch is not None else bool(op_defaults.get("auto_close_on_mismatch", False))
    if not terminal_path:
        raise HTTPException(status_code=400, detail="MT5_TERMINAL_PATH es obligatorio")
    if int(mt5_login) <= 0:
        raise HTTPException(status_code=400, detail="MT5_LOGIN debe ser un entero positivo")
    if not mt5_password:
        raise HTTPException(status_code=400, detail="MT5_PASSWORD es obligatorio")
    if not mt5_server:
        raise HTTPException(status_code=400, detail="MT5_SERVER es obligatorio")
    if total_volume <= 0:
        raise HTTPException(status_code=400, detail="TOTAL_VOLUME debe ser mayor a 0")
    if near_entry_pips_min < 0:
        raise HTTPException(status_code=400, detail="NEAR_ENTRY_PIPS_MIN no puede ser negativo")
    if near_entry_spread_mult < 0:
        raise HTTPException(status_code=400, detail="NEAR_ENTRY_SPREAD_MULT no puede ser negativo")

    env = os.environ.copy()
    env["PYTHONUNBUFFERED"] = "1"
    env["MT5_TERMINAL_PATH"] = terminal_path
    env["MT5_LOGIN"] = str(mt5_login)
    env["MT5_PASSWORD"] = mt5_password
    env["MT5_SERVER"] = mt5_server
    env["EXECUTION_MODE"] = "single"
    env["EXECUTION_PROFILE"] = execution_profile
    env["TOTAL_VOLUME"] = str(total_volume)
    env["NEAR_ENTRY_PIPS_MIN"] = str(near_entry_pips_min)
    env["NEAR_ENTRY_SPREAD_MULT"] = str(near_entry_spread_mult)
    env["VERIFY_ORDER_AFTER_SEND"] = "true" if verify_after_send else "false"
    env["AUTO_CLOSE_ON_MISMATCH"] = "true" if auto_close_mismatch else "false"
    env["TRADING_BOT_DB_PATH"] = str(DB_PATH)

    try:
        operador_manager.start(env)
        _persist_process_runtime_env(PROCESS_NAME_OPERADOR, env)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return {"status": "started"}


@app.post("/api/stop/operador")
def stop_operador():
    operador_manager.stop()
    return {"status": "stopped"}


@app.get("/api/logs/lector")
def logs_lector():
    return StreamingResponse(_sse_stream(lector_manager.log), media_type="text/event-stream")


@app.get("/api/logs/operador")
def logs_operador():
    return StreamingResponse(_sse_stream(operador_manager.log), media_type="text/event-stream")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=8000, log_level="warning", access_log=False)

